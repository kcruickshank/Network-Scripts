import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from pandas import DataFrame
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import numpy as np
import re
from re import search
import os
import shutil
import sys
import netmiko
from netmiko import ConnectHandler
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
import clean_int_status as clean_status
import clean_mac as clean_mac
import clean_int_desc as clean_desc
import clean_arp as clean_arp
import getpass
import datetime as datetime

# Define some variable to be used later in the script and ask user for some input
username = input('Enter Username: ')
password = getpass.getpass('Enter Password: ')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Switch_Audit_Files\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Switch_Audit_Files\\Final_Output_Files\\'
tftp_path_move_old = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Switch_Audit_Files\\Old_Text_Files\\'
number_connected = 0
number_notconnect = 0
number_disabled = 0
number_err_disabled = 0
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y_%H-%M")  # Set timestamp to current system time
logfile = timestamp + '_Switch_Audit_Error_Log.txt'

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)
count_ip = df_read_ip.shape[0]  # Get the number of rows in column 1
if count_ip == 1:
    print("Starting to get files from " + str(count_ip) + " switch")
else:
    print("Starting to get files from " + str(count_ip) + " switches")

connect_success = False
connect_issue = False
for index, row in df_read_ip.iterrows():
    ip_address = df_read_ip.iloc[index, 0]  # Set IP Address Variable
    # Define a switch type
    switch = {
        "device_type": "cisco_ios",
                    "ip": ip_address,
                    "username": username,
                    "password": password,
    }

    try:
        ############################################
        # Connect to switch and put in enable mode #
        ############################################
        net_connect = ConnectHandler(**switch)
        net_connect.enable()
        
        # Turn on shell processing on switch so can use linux uname command
        # Maybe future as not all switches support it
        #net_connect.send_config_set("shell processing full")  
        switchname = net_connect.send_command ("sh ver | i uptime")  # Use to get the hostname
        switchname = switchname.split()[0]  # Get the first word which will be the switchname
        #switchname = switchname.strip()  # Strip any trailing white space from the variable

        # Assign IOS commands to variables
        command_1 = "sh mac address-table | redirect tftp://" + str(ip_tftp_server) + "/Switch_Audit_Files/" + switchname + "_mac_add.txt"
        command_2 = "sh int status | redirect tftp://" + str(ip_tftp_server) + "/Switch_Audit_Files/" + switchname + "_int_status.txt"
        command_3 = "sh int desc | redirect tftp://" + str(ip_tftp_server) + "/Switch_Audit_Files/" + switchname + "_int_desc.txt"
        command_4 = "sh arp | redirect tftp://" + str(ip_tftp_server) + "/Switch_Audit_Files/" + switchname + "_arp.txt"

        print("Copying files for " + switchname + " to TFTP Server " + ip_tftp_server + "...")
        #print(" ")
        #Send the commands to the switch
        net_connect.send_command(command_1)
        net_connect.send_command(command_2)
        net_connect.send_command(command_3)
        net_connect.send_command(command_4)
        net_connect.disconnect()  # Disconnect from switch
        print("Finished copying files for " + switchname + ".")
        connect_success = True
    except (AuthenticationException):
        print ('Authentication Failure: ' + ip_address)
        Authfailure=open(tftp_path + str(logfile), "a")
        Authfailure.write('\nAuthentication Failure for IP ' + ip_address)
        Authfailure.close()
        connect_issue = True
        continue 
    except (NetMikoTimeoutException):
        print ('\n' + 'Timeout to device: ' + ip_address)
        Timeouts=open(tftp_path + str(logfile), "a")
        Timeouts.write('\nTimeout to device: ' + ip_address)
        Timeouts.close()
        connect_issue = True
        continue
    except (SSHException):
        print ('SSH might not be enabled: ' + ip_address)
        SSHException=open(tftp_path + str(logfile), "a")
        SSHException.write('\nSSH might not be enabled: ' + ip_address)
        SSHException.close()
        connect_issue = True
        continue 
    except (EOFError):
        print ('\n' + 'End of liner error attempting device: ' + ip_address)
        EOFError=open(tftp_path + str(logfile), "a")
        EOFError.write('\nEnd of liner error attempting device: ' + ip_address)
        EOFError.close()
        connect_issue = True
        continue
    except unknown_error:
        print ('Some other error: ' + str(unknown_error))
        UnknownError=open(tftp_path + str(logfile), "a")
        UnknownError.write('\nSome other error: ' + str(unknown_error) + ' for ' + ip_address)
        UnknownError.close()
        connect_issue = True
        continue

# Check to see if some files have been written to TFTP Directory
if connect_success != True:
    print("\nThere were issues getting files, please review log file " + tftp_path + logfile + "\n")
    sys.exit()
else:
    if connect_issue == True:
        print("\nFinished TFTP Transfers, but there were some errors\n")
        print("Please review log file for errors " + tftp_path + logfile + "\n")
        # Prompt for the name of the output spreadsheet
        spreadsheet_name = input("What name do you want to save the spreadsheet as: ")
        spreadsheet_name = spreadsheet_name.strip()  # Strip any unwanted whitespace at end
    else:
        print("\nFinished TFTP Transfers, with no errors\n")
        # Prompt for the name of the output spreadsheet
        spreadsheet_name = input("What name do you want to save the spreadsheet as: ")
        spreadsheet_name = spreadsheet_name.strip()  # Strip any unwanted whitespace at end

    # Create Spreadsheet with blank sheet
    writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name + '.xlsx', engine='xlsxwriter')
    writer.save()

    # Assign the new spreadsheet to Pandas Excel Writer using
    # openpyxl engine in append mode 
    #writer = pd.ExcelWriter(tftp_path + spreadsheet_name + '.xlsx', engine='openpyxl', mode='a')

    ###############################################################################
    # Create a new DataFrame and then read through the switch audit directory,    #
    # create a list of the switch names and then run the code against each        #
    ###############################################################################
    # Create Blank Dataframe for the final arp table
    blank_arp_cols = ['MAC Address', 'IP Address']
    blank_arp_data = [[' ', ' ']]
    dfarp = pd.DataFrame(blank_arp_data, columns=blank_arp_cols)

    df0 = pd.DataFrame()
    filenames = os.listdir(tftp_path)
    os.chdir(tftp_path)
    for filename in filenames:  # Convert the text file names to lowercase
        if '.txt' in filename:
            newfilename = filename.lower()
            os.rename(filename, newfilename)
    df0['FILES'] = os.listdir(tftp_path)
    df0 = df0.loc[~df0['FILES'].str.contains('Files', flags=re.I, regex=True)]
    df0 = df0.loc[~df0['FILES'].str.contains('.xlsx', flags=re.I, regex=True)]
    df0 = df0.loc[~df0['FILES'].str.contains('.csv', flags=re.I, regex=True)]
    df0 = df0.loc[~df0['FILES'].str.contains('error_log', flags=re.I, regex=True)]
    full_list_files = df0['FILES'].tolist()
    df0 = df0.replace({'_int_desc.txt': ''}, regex=True)
    df0 = df0.replace({'_int_status.txt': ''}, regex=True)
    df0 = df0.replace({'_mac_add.txt': ''}, regex=True)
    df0 = df0.replace({'_arp.txt': ''}, regex=True)
    df0 = df0.drop_duplicates()
    df0.reset_index(drop=True, inplace=True)
    file_list = df0['FILES'].tolist()

    # Print no files if file list is blank
    if file_list == []:
        print("There are no files to work on, please get files from switch.")
        sys.exit()  # Exit the script

    # Check to see if DNS.csv file exists
    dns_file_path = tftp_path_files + 'DNS.csv'
    dns_file_exist = os.path.isfile(dns_file_path)

    if dns_file_exist == False:
        print("\n")
        print("The DNS.csv file is missing from " + tftp_path + "\n")
        print("Do you want to continue without DNS file, y or n? ")
        answer = input(': ')
        answer = answer.lower()
        answered = False
        while answered != True:
            print(answer)
            if answer == 'y':
                answered = True
            elif answer == 'n':
                print("Scipt ending, please copy DNS.csv to " + tftp_path)
                sys.exit()  # Exit the script
            else:
                print("Invalind answer, accepts only letter y or n. \n")
                answer = input(': ')
                answer = answer.lower()
    elif dns_file_exist == True:  
        # Read the csv DNS file
        df6 = pd.read_csv(str(tftp_path_files) + 'DNS.csv', header=None)
        df6 = pd.DataFrame(df6)
        df6 = df6.drop([0], axis=0)  # Drop row index 0 that will have the headings read in from the text file
        dns_columns = ["IP Address", "DNS Entry"]
        df6.columns = dns_columns
        writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name + '.xlsx', engine='openpyxl', mode='a')
        # Save the Dataframe to a new sheet called DNS
        df6.to_excel(writer, sheet_name='DNS', startrow=0, index=None)
        writer.save()  # Save the Spreadsheet
        df6 = pd.DataFrame()  # Clear Dataframe

    # Start looping through the switch names and running the code against each.
    for file in file_list:
        arp = False
        arp_firewall = False
        arp_spreadsheet = False
        arp_fwspreadsheet = False
        int_desc = False
        int_status = False
        mac_add = False
        df1 = pd.DataFrame()  # Clear Dataframe
        df2 = pd.DataFrame()  # Clear Dataframe
        df3 = pd.DataFrame()  # Clear Dataframe
        df4 = pd.DataFrame()  # Clear Dataframe
        df5 = pd.DataFrame()  # Clear Dataframe
        df1_list = []  # Clear List
        df2_list = []  # Clear List
        df4_length = 0
        switchname = str(file).strip()
        sub_string = "fiw"
        # Check if the particular files exist
        for filename in full_list_files:
            #print(filename)
            if filename == str(switchname) + "_arp.txt":
                arp = True
                if search(sub_string, switchname):
                    arp_firewall = True       
            if filename == str(switchname) + "_int_desc.txt":
                int_desc = True
            if filename == str(switchname) + "_int_status.txt":
                int_status = True
            if filename == str(switchname) + "_mac_add.txt":
                mac_add = True
        
        if int_desc == True:
            ########################################################################################
            # Start Data Frame 1 and read in the contents from the interface description text file #
            ########################################################################################

            # Call clean method for interface description text file
            clean_desc.clean_file(switchname, tftp_path)

            # Read the text file that has been cleaned and sorted into columns
            df1 = pd.read_csv(str(tftp_path) + switchname + '_int_desc_clean.csv', header=None)
            df1 = pd.DataFrame(df1)  # Set df1 as type DataFrame
            df1 = df1.drop([0, 1], axis=0)  # Drop row index 0 that will have the headings read in from the text file
            df1 = df1.drop(columns=[1, 2], axis=1)  # Drop the unwanted columns 2 and 3
            df1.to_csv(str(tftp_path) + switchname + '_int_desc_clean.csv', index=False, header=None)  # Save file out with dropped columns
            df1 = pd.read_csv(str(tftp_path) + switchname + '_int_desc_clean.csv', header=None)  # Read in the ammended file
            df1 = pd.DataFrame(df1)  # Set df1 as type DataFrame
            df1_length = len(df1.axes[1])  # Get the number of Columns

            desc_columns = ["Interface", "Description"]  # Set the default columns
            i = 2
            if df1_length > 2: # If there are more columns than the default, pad out with None Columns
                i = 2
                while i < df1_length:
                    desc_columns.append('None' + str(i))
                    i += 1
                df1.columns = desc_columns  # Create our own header row
                i = 2  # Reset variable i
                while i < df1_length:
                    df1['None' + str(i)] = df1['None' + str(i)].replace(np.nan, " ")  # Replace Non Values with blank
                    df1['None' + str(i)] = df1['None' + str(i)].astype(str)  # Convert the dataframe columns to a type string
                    df1['Description'] = df1['Description'] + ' ' + df1['None' + str(i)]  # Loop through the None columns and add the string data to the Description
                    i += 1
                i = 2  # Reset variable i
                while i < df1_length:  # Loop through and drop the extra columns
                    df1 = df1.drop(columns=['None' + str(i)])
                    i += 1
            else:
                df1.columns = desc_columns  # Create our own header row

            df1['Description'] = df1['Description'].replace(np.nan, " ")  # Replace Non Values with blank
            df1 = df1[df1['Interface'].notna()]  # Remove the rows that have Nan value in the Interface Column
            # Use regular expression to filter out unwanted rows that match string
            df1 = df1.loc[~df1['Interface'].str.contains('Vl|Tu|Lo', flags=re.I, regex=True)]
            df1.reset_index(drop=True, inplace=True)  # Reset the indexes

            # Clean up files and print finised editing
            os.remove(str(tftp_path) + switchname + '_int_desc_clean.csv')
            shutil.move(str(tftp_path) + switchname + '_int_desc.txt', str(tftp_path_move_old) + switchname + '_int_desc.txt')
            print("Finished editing INT Desc text file for " + str(switchname))

        if int_status == True:
            ###################################################################################
            # Start Data Frame 2 and read in the contents from the interface status text file #
            ###################################################################################

            clean_status.clean_file(switchname, tftp_path)

            df2 = pd.read_csv(str(tftp_path) + switchname + '_int_status_clean.csv', header=None)
            df2 = pd.DataFrame(df2)
            df2_length = len(df2.axes[1]) # Get the number of Columns
            status_columns = ["Port", "Port Description", "Status", "Vlan/Trunk", "Duplex", "Speed", "Port Type"]  # Define the column names

            df2 = df2.drop(0)  # Drop row index 0 that had the extra Headings from the clean function
            df2 = df2.drop(1)  # Drop row index 1 that had the top line with the heading names read from the text file
            df2.reset_index(drop=True, inplace=True)  # Reset the row indexes of the DataFrame

            i = 7  # Set variable i to the expected number of columns
            if df2_length > 7: # If there are more columns than the default, pad out with None Columns
                while i < df2_length:
                    status_columns.append("None" + str(i))
                    i += 1
                df2.columns = status_columns  # Assign the column names to the DataFrame
                i = 7  # Reset variable i to start from the first new column
                while i < df2_length:
                    df2['None' + str(i)] = df2['None' + str(i)].replace(np.nan, " ")  # Replace Non Values with blank
                    df2['None' + str(i)] = df2['None' + str(i)].astype(str)  # Convert the dataframe columns to a type string
                    df2['Port Type'] = df2['Port Type'] + ' ' + df2['None' + str(i)]  # Loop through the None columns and add the string data to the Type Column
                    i += 1
                i = 7  # Reset variable i
                while i < df2_length:  # Loop through and drop the extra columns
                    df2 = df2.drop(columns=['None' + str(i)])
                    i += 1
            else:
                df2.columns = status_columns  # Use the default header row with 7 columns

            df2 = df2[df2['Port'].notna()]  # Remove the rows that have Nan value in the Port Column
            df2.reset_index(drop=True, inplace=True)  # Reset the indexes so we can drop the 0 index which will the old header row

            ####################################################################################
            # Make sure that the Interface Status Text and Interface Description text files    #
            # match on the interfaces. Remove any interfaces that are found in the description #
            # file that don't match in the status text file                                    #
            ####################################################################################

            # Convert columns to lists
            df1_list = df1['Interface'].tolist()
            df2_list = df2['Port'].tolist()

            # Loop through lists and write out the matching and non matching values
            matching_vals = []
            no_match = []
            for val in df1_list:
                if val in df2_list:
                    matching_vals.append(val)
                else:
                    if val == " ":
                        continue
                    else:
                        no_match.append(val)

            str_no_match = '|'.join(str(e) for e in no_match)  # Converts list array to a string value

            if str_no_match != "":  # If there are values in the no matches then remove the rows from df1
                df1 = df1.loc[~df1['Interface'].str.contains(str_no_match, flags=re.I, regex=True)]
                df1.reset_index(drop=True, inplace=True)  # Reset the indexes

            # Lets sort the interfaces in DF2 to be in the same order as DF1, so that when we come to compare and grab the
            # description that we don't get an error. Example if gig interfaces some switches have the Fa0 port. Int description
            # output puts it at the top and int status puts it at the bottom
            df2['Port'] = pd.Categorical(df2['Port'], matching_vals)  # Apply how the list should be sorted to the Ports Column
            df2 = df2.sort_values("Port")  # Now sort the column
            df2.reset_index(drop=True, inplace=True)  # Reset the indexes

            # Get the interface description from Data Set 1 by matching the interfaces between the dataframes
            df2['Port Description'] = np.where(df1['Interface'] == df2['Port'], df1['Description'], 'No Match')

            
            # Copy DF2 to new Dataframe DF6, then manipulate DF6 ready to count the interfaces
            df6 = df2
            # Remove any rows with PO in the interfaces as we don't want to count the portchannels
            df6 = df6.loc[~df6['Port'].str.contains('Po', flags=re.I, regex=True)]  

            # Create a list from the status column and then coun
            df2_list_status = df6['Status'].tolist()
            df2_list_status_lwr = []

            # Just to make sure convert values in list to lowercase
            for item in df2_list_status:
                df2_list_status_lwr.append(item.lower())

            number_connected = df2_list_status_lwr.count("connected")
            number_disabled = df2_list_status_lwr.count("disabled")
            number_notconnect = df2_list_status_lwr.count("notconnect")
            number_err_disabled = df2_list_status_lwr.count("err-disabled")
            
            # Create a new CSV file and drop the index
            #df2.to_csv(str(tftp_path_final) + switchname + '_interface_final.csv', index=False)

            # Clean up files and print finised editing
            os.remove(str(tftp_path) + switchname + '_int_status_clean.csv')
            shutil.move(str(tftp_path) + switchname + '_int_status.txt', str(tftp_path_move_old) + switchname +
                '_int_status.txt')
            print("Finished editing INT Status text file for " + str(switchname))

        if mac_add == True:
            ###################################################################################
            # Start Data Frame 3 and 4 to read in the contents from the MAC Address text file #
            # Sort the Interface Value column                                                 #
            ###################################################################################

            #Read the text file using Pandas fixed width formatted lines
            df3 = pd.read_fwf(str(tftp_path) + switchname + '_mac_add.txt', header=None)
            df3 = pd.DataFrame(df3)

            # Use regular expression to filter out unwanted rows
            df3 = df3.loc[~df3[0].str.contains('--|vlan|mac|all|unicast|system|multicast|switch|duplicate|Legend|age|entry|'
                                            'supervisor|remove|n/a|vl', flags=re.I, regex=True)]
            df3 = df3.replace({'Port-channel': 'Po'}, regex=True)  # Replace Port-channel with po
            df3 = df3.replace({'GigabitEthernet': 'Gi'},
                            regex=True)  # Replace GigabitEthernet with Gi so it matches on lookup with interfaces output file
            df3 = df3.replace({'ip': ''}, regex=True)  # Remove word starts with ip
            df3 = df3.replace({'x': ''},
                            regex=True)  # There was the word ipx, ip was removed in previous now we need to remove the x left over
            df3 = df3.replace({'\*': ''}, regex=True)  # Remove asterisk
            df3 = df3.replace({'assigned': ''}, regex=True)  # Remove word assigned
            df3 = df3.replace({'other': ''}, regex=True)  # Remove word other
            df3 = df3.replace({'Yes': ''}, regex=True)  # Remove unwanted data
            df3 = df3.replace({'No': ''}, regex=True)  # Remove unwanted data
            # If the line does not contain dynamic then remove the row
            df3 = df3.loc[df3[0].str.contains('dynamic', flags=re.I, regex=True)]  # Remove rows that don't have Dynamic
            # Save Temp file out that will be opened by the clean function
            df3.to_csv(str(tftp_path) + switchname + '_mac_add_temp1.txt', index=False, header=None, quoting=3, escapechar=",")

            clean_mac.clean_file(switchname, tftp_path)  # Call function to clean of Mac address text file

            # Now work with the amended temp file that will now have the text file with the required 4 columns
            df4 = pd.read_csv(str(tftp_path) + switchname + '_mac_add_clean.csv', header=None)  # Open the cleaned Mac Temp File
            df4 = pd.DataFrame(df4)  # Set it as type DataFrame
            df4_length = len(df4.axes[1])  # Get the number of Columns
            df4_col_max_index = df4_length - 1  # Set the max column index
            mac_columns = ["Vlan", "Mac Address", "Type", "Port"]  # Set the default columns
            df4 = df4.astype(str)
            pattern_list = ['dynamic', 'DYNAMIC']  # Pattern for column 2
            pattern_list1 = ['Po', 'Fa', 'Gi', 'Te']  # Pattern for column 3
            pattern_type = '|'.join(pattern_list)  # Convert into string
            pattern_ports = '|'.join(pattern_list1)  # Convert into string

            # The following coded loops through columns 2 and 3 to ensure they include the correct data
            # based on the patterns above
            i = 2
            while i <= df4_col_max_index:
                if df4[2].str.contains(pattern_type).any() == True:
                    i = 100
                else:
                    df4 = df4.drop(columns=[i], axis=1)
                    i += 1

            i = 3
            while i <= df4_col_max_index:
                if df4[i].str.contains(pattern_ports).any() == True:
                    i = 100
                else:
                    df4 = df4.drop(columns=[i], axis=1)
                    i += 1

            df4.columns = mac_columns  # Assign the column names to the DataFrame
            df4 = df4.drop(['Type'], axis=1)  # Drop the unwanted Type column

            # Define how the Ports Column has to be ordered based on the interface list from DF1
            df4['Port'] = pd.Categorical(df4['Port'], matching_vals)  # Apply how the list should be sorted to the Ports Column
            df4 = df4.sort_values("Port")  # Sort the Ports column
            df4.reset_index(drop=True, inplace=True)  # Reset Index
            # Add in the missing interfaces
            df4["All_Ports"] = df1['Interface']
            df4['Not_in_List'] = np.where(df4['All_Ports'].isin(df4['Port']), " ", df4['All_Ports'])
            df4['Not_in_List'] = df4['Not_in_List'].replace(np.nan, " ")  # Remove the blanks from the columns
            new_list = df4['Not_in_List']  # Write the column to a list
            xList = list(set(new_list)-set(df4['Port']))  # Create list variable of list and existing Ports column
            df4 = df4.append(pd.DataFrame({'Port': xList}), ignore_index=True)  # Append the list
            df4['Port'] = pd.Categorical(df4['Port'], matching_vals)  # Apply how the list should be sorted to the Ports Column
            df4 = df4.sort_values("Port") # Now sort the column
            df4 = df4.drop(['All_Ports', 'Not_in_List'], axis=1)  # Drop the temp columns before writing to file
            df4.reset_index(drop=True, inplace=True)  # Reset Index
            #df4.to_csv(str(tftp_path_final) + switchname + '_mac_add_final.csv', index=False)  # Save amended file
            #with pd.ExcelWriter('C:\TFTP-Root\\Switch_Audit.xlsx', engine="openpyxl", mode='a') as writer:
            #    df4.to_excel(writer, sheet_name=str(switchname), index=False)
            # Clean up files and print finised editing
            os.remove(str(tftp_path) + switchname + '_mac_add_clean.csv')
            os.remove(str(tftp_path) + switchname + '_mac_add_temp1.txt')
            shutil.move(str(tftp_path) + switchname + '_mac_add.txt', str(tftp_path_move_old) + switchname + '_mac_add.txt')
            print("Finished editing MAC Address text file for " + str(switchname))
        
        if arp == True:

            ########################################################################################
            # Start Data Frame 5 depending on if its a firewall or core switch do different things #
            ########################################################################################
            arp_columns = ["IP Address", "MAC Address"]  # Define the column names
            column_swap = ["MAC Address", "IP Address"]  # Define how the columns need to be ordered
            if arp_firewall == True:
                # Call the clean arp function
                clean_arp.clean_file(switchname,tftp_path)
                
                # Read the csv arp file
                df5 = pd.read_csv(str(tftp_path) + switchname + '_arp_clean.csv', header=None)
                df5 = df5.drop(columns=[0, 3], axis=1)
                df5 = df5.loc[~df5[2].str.contains('Incomplete', flags=re.I, regex=True)]
                df5.columns = arp_columns  # Assign the column names to the DataFrame
                df5 = df5.reindex(columns=column_swap)
                #df5.to_csv(str(tftp_path_final) + switchname + '_final_arp.csv', index=False, quoting=3, escapechar=",")
                arp_fwspreadsheet = True
                dfarp = pd.concat([dfarp, df5], axis=0)
                # Clean up files and print finised editing
                os.remove(str(tftp_path) + switchname + '_arp_clean.csv')
                shutil.move(str(tftp_path) + switchname + '_arp.txt', str(tftp_path_move_old) + switchname + '_arp.txt')
                print("Finished editing ARP text file for " + str(switchname))
            else:
                # Call the clean arp function
                clean_arp.clean_file(switchname,tftp_path)

                # Read the csv arp file 
                df5 = pd.read_csv(str(tftp_path) + switchname + '_arp_clean.csv', header=None)
                df5 = df5.drop(columns=[0, 2, 4, 5], axis=1)
                df5 = df5.loc[~df5[3].str.contains('Incomplete', flags=re.I, regex=True)]
                df5.columns = arp_columns  # Assign the column names to the DataFrame
                df5 = df5.reindex(columns=column_swap)
                #df5.to_csv(str(tftp_path_final) + switchname + '_final_arp.csv', index=False, quoting=3, escapechar=",")
                #arp_spreadsheet = True
                dfarp = pd.concat([dfarp, df5], axis=0)

                # Clean up files and print finised editing
                os.remove(str(tftp_path) + switchname + '_arp_clean.csv')
                shutil.move(str(tftp_path) + switchname + '_arp.txt', str(tftp_path_move_old) + switchname + '_arp.txt')
                print("Finished editing ARP text file for " + str(switchname))
        
        # After running the above we can now combine the values into one dataframe and write 
        # to an excel file and format.
        df4['IP Address'] = ''
        df4['DNS'] = ''
        df4 = pd.merge(df4, df2)
        print("Finishied combining values to DataFrame for " + str(switchname))

        # Start to build the spreadsheet
        print('Building worksheet for ' + str(switchname + '...'))
        if arp_fwspreadsheet == True:
            continue
        else:
            # Assign the new spreadsheet to Pandas Excel Writer using
            # openpyxl engine in append mode 
            writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name + '.xlsx', engine='openpyxl', mode='a')

            # Save the Dataframe to a new sheet called the switchname variable
            df4.to_excel(writer, sheet_name=switchname, startrow=6, index=None)
            writer.save()  # Save the Spreadsheet

            # Load the workbook to get all sheets
            workbook = load_workbook(tftp_path_final + spreadsheet_name + '.xlsx')  # Update the workbook
                
            worksheet=workbook[switchname]  # Set the active worksheet
            worksheet.sheet_view.showGridLines = False  # Hide gridlines on sheet

            # Freeze the Header Row
            freeze_row = worksheet['A8']
            worksheet.freeze_panes = freeze_row

            # Setup some colours for some cells
            light_yellow = 'ebeca4'
            dark_blue = '0c4672'
            heading_green = '88c184'
            #subheading_orange = 'e58443'
            subheading_orange = 'f4c4a4'
            white_font = 'ffffff'
            
            # Set Heading Colours
            worksheet['A7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['B7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['C7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['D7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['E7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')    
            worksheet['F7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['G7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['H7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['I7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['J7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            worksheet['K7'].fill = PatternFill(fgColor=heading_green, fill_type='solid')

            # Add some text to the top of the spreadsheet above the rows
            # of data from the dataframe
            wcell1 = worksheet.cell(2,1)  # Set cell to A2
            wcell1.value = "Switch Audit for " + switchname  # Text for the what has been audited
            worksheet['A2'].font = Font(bold=True)  # Bold the Text
            worksheet['A2'].fill = PatternFill(fgColor=light_yellow, fill_type='solid')
            worksheet.merge_cells('A2:C2')  # Merge the cells A2 to C2
            worksheet['A2'].alignment = Alignment(horizontal='center')
            wcell2 = worksheet.cell(2,6)
            wcell2.value = "Number of Ports Connected: " + str(number_connected)
            worksheet['F2'].font = Font(bold=True, color=white_font)
            worksheet['F2'].fill = PatternFill(fgColor=dark_blue, fill_type='solid')
            wcell3 = worksheet.cell(3,6)
            wcell3.value = "Number of Ports Not Connected: " + str(number_notconnect)
            worksheet['F3'].font = Font(bold=True, color=white_font)
            worksheet['F3'].fill = PatternFill(fgColor=dark_blue, fill_type='solid')
            wcell4 = worksheet.cell(4,6)
            wcell4.value = "Number of Ports Disabled: " + str(number_disabled)
            worksheet['F4'].font = Font(bold=True, color=white_font)
            worksheet['F4'].fill = PatternFill(fgColor=dark_blue, fill_type='solid')
            #worksheet.merge_cells('F2:H2')
            #worksheet.merge_cells('F3:H3')
            #worksheet.merge_cells('F4:H4')

            # Add some headings above the header to explain where the data has came from
            wcell5 = worksheet.cell(6,1)  # Set cell to A6
            wcell5.value = "From MAC Address Table"  # Text for the what has been audited
            worksheet['A6'].font = Font(bold=True)  # Bold the Text
            worksheet['A6'].fill = PatternFill(fgColor=subheading_orange, fill_type='solid')
            worksheet.merge_cells('A6:C6')  # Merge the cells A6 to C6
            worksheet['A6'].alignment = Alignment(horizontal='center')

            wcell6 = worksheet.cell(6,6)  # Set cell to F6
            wcell6.value = "From Interface Status and Description"  # Text for the what has been audited
            worksheet['F6'].font = Font(bold=True)  # Bold the Text
            worksheet['F6'].fill = PatternFill(fgColor=subheading_orange, fill_type='solid')
            worksheet.merge_cells('F6:K6')  # Merge the cells F6 to K6
            worksheet['F6'].alignment = Alignment(horizontal='center')

            # Highlight cells that contain particular text by using a special formula
            # Use this website to get colours https://www.color-hex.com/color/b8e3bd
            # Define the different colours
            red_text = Font(color="9C0006")
            red_fill = PatternFill(bgColor="FFC7CE")
            dxf1 = DifferentialStyle(font=red_text, fill=red_fill)

            green_text = Font(color="1f2b21")
            green_fill = PatternFill(bgColor="b8e3bd")
            dxf2 = DifferentialStyle(font=green_text, fill=green_fill)

            # Setup the Conditional Statement Rules
            rule1 = Rule(type="containsText", operator="containsText", text="notconnect", dxf=dxf1)
            rule1.formula = ['NOT(ISERROR(SEARCH("notconnect",G8)))']
            worksheet.conditional_formatting.add('G8:G1000', rule1)

            rule2 = Rule(type="containsText", operator="containsText", text="disabled", dxf=dxf1)
            rule2.formula = ['NOT(ISERROR(SEARCH("disabled",G8)))']
            worksheet.conditional_formatting.add('G8:G1000', rule2)

            rule3 = Rule(type="containsText", operator="containsText", text="err-disabled", dxf=dxf1)
            rule3.formula = ['NOT(ISERROR(SEARCH("err-disabled",G8)))']
            worksheet.conditional_formatting.add('G8:G1000', rule3)

            rule4 = Rule(type="containsText", operator="containsText", text="connected", dxf=dxf2)
            rule4.formula = ['NOT(ISERROR(SEARCH("connected",G8)))']
            worksheet.conditional_formatting.add('G8:G1000', rule4)

            # Set Min and Max Rows
            min_row = 8
            max_row = worksheet.max_row

            # Add some borders
            thin_border = Border(
                left=Side(border_style=BORDER_THIN), 
                right=Side(border_style=BORDER_THIN), 
                top=Side(border_style=BORDER_THIN), 
                bottom=Side(border_style=BORDER_THIN)
                )
            
            i = 8
            rows2 = worksheet.iter_rows(min_row,max_row)
            for row2 in rows2:
                row2[3].value = '=IFERROR(VLOOKUP(B' + str(i) + ',ARP!A:B,2,FALSE), "Not found")'
                row2[4].value = '=IFERROR(VLOOKUP(D' + str(i) + ',DNS!A:B,2,FALSE), "Not found")'
                i += 1
                for cell in row2:
                    cell.border = thin_border 

            worksheet.cell(2,1).border = thin_border
            worksheet.cell(2,2).border = thin_border
            worksheet.cell(2,3).border = thin_border

            worksheet.cell(2,6).border = thin_border
            worksheet.cell(3,6).border = thin_border
            worksheet.cell(4,6).border = thin_border

            worksheet.cell(6,1).border = thin_border
            worksheet.cell(6,2).border = thin_border
            worksheet.cell(6,3).border = thin_border

            worksheet.cell(6,6).border = thin_border
            worksheet.cell(6,7).border = thin_border
            worksheet.cell(6,8).border = thin_border
            worksheet.cell(6,9).border = thin_border
            worksheet.cell(6,10).border = thin_border
            worksheet.cell(6,11).border = thin_border

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 5
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 10
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 17
            worksheet.column_dimensions['F'].width = 65
            worksheet.column_dimensions['G'].width = 10
            worksheet.column_dimensions['H'].width = 10
            worksheet.column_dimensions['I'].width = 8
            worksheet.column_dimensions['J'].width = 9
            worksheet.column_dimensions['K'].width = 25

            # Save the workbook
            workbook.save(tftp_path_final + spreadsheet_name + '.xlsx')

    # Write the final ARP sheet to the workbook
    writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name + '.xlsx', engine='openpyxl', mode='a')
    dfarp = dfarp.drop(0)  # Drop the dummy Blank row inserted at the start of script
    dfarp = dfarp.drop_duplicates()  # Drop Duplicates
    dfarp.to_excel(writer, sheet_name='ARP', startrow=0, index=None)  # Write Final ARP Sheet
    writer.save()  # Save the Spreadsheet
    # Load the workbook to get all sheets
    workbook = load_workbook(tftp_path_final + spreadsheet_name + '.xlsx')
    # Delete the blank sheet
    workbook.remove(workbook['Sheet1'])
    # Move ARP sheet to the beginning of the workbook
    sheets = workbook._sheets
    from_loc = len(sheets) - 1
    to_loc = 0
    arp_sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, arp_sheet) 
    # Save final changes to the workbook
    print("Finished creating Spreadsheet " + spreadsheet_name + ".")
    workbook.save(tftp_path_final + spreadsheet_name + '.xlsx')