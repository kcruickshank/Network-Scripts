import pandas as pd
from pandas import ExcelWriter
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading
import clean_file as clean_file

# These capture errors relating to hitting ctrl+C (I forget the source)
#signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # IOError: Broken pipe
#signal.signal(signal.SIGINT, signal.SIG_DFL)  # KeyboardInterrupt: Ctrl-C

# Set the number of threads, I've found that 5 is the max
num_threads = 10

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()

# Define some variables to be used later in the script and ask user for some input
#username = input('Enter Username: ')
username = 'adm.cruickshank'
password = getpass.getpass('Enter Password: ')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Wireless\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Wireless\\'
connection_details = []

# Define the current date and error log file info
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y")  # Set timestamp to current system date
timestamp1 = now.strftime("%d/%m/%Y at %H:%M")  # Set timestamp to current system time including hour and minutes
logfile = timestamp + '_error_connection_log.txt'
tempoutput1 = '_sysinfo.txt'
tempoutput2 = 'WLC Radius_Summary.txt'
tempoutput3 = '_Wlan_Summary.txt'
tempoutput4 = '_Summary.txt'
tempoutput5 = '_WLANs_RADIUS_Summary.txt'
spreadsheet_name = timestamp + '_WLC_Check.xlsx'

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)
#count_ip = df_read_ip.shape[0]  # Get the number of rows in column 1

# Create Spreadsheet with blank sheet
#writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name, engine='xlsxwriter')
#writer.save()
# Load the workbook
#workbook = load_workbook(tftp_path_final + spreadsheet_name)


check_file_path_1 = tftp_path + 'WLC Radius_Summary.txt'
file_exist_1 = os.path.isfile(check_file_path_1)

if file_exist_1:
    os.remove(check_file_path_1)


# Start to iterate through the IP's in the file
def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()
        print("{}: Acquired IP: {}".format(i,ip_address))

        # Define a switch type
        wlc = {
            "device_type": "cisco_wlc",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
        }

        # Test the ssh connection and handle any errors and output to text file
        try:
            net_connect = ConnectHandler(**wlc)
        except (AuthenticationException):
            with print_lock:
                print("\n{}: ERROR: Authentication failed for {}. Stopping thread. \n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: Authentication failed for {} on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            # Add connection failed to list
            connection_details.append("Failed")
            q.task_done()
            continue 
        except (NetMikoTimeoutException):
            with print_lock:
                print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: Connection to {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            # Add connection failed to list
            connection_details.append("Failed")
            q.task_done()
            continue
        except (SSHException):
            with print_lock:
                print("\n{}: SSH might not be enable on: {} timed-out.\n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: SSH might not be enabled on: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            # Add connection failed to list
            connection_details.append("Failed")
            q.task_done()
            continue
        except (EOFError):
            with print_lock:
                print("\n{}: End of liner error attempting device: {} timed-out.\n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: End of line error attempting device: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            # Add connection failed to list
            connection_details.append("Failed")
            q.task_done()
            continue


        # Check to see if the WLC Radius Summary text file exists, if it does then 
        # delete it so a new one can be created

        check_file_path_2 = tftp_path + ip_address + str(tempoutput5)
        file_exist_2 = os.path.isfile(check_file_path_2)

        if file_exist_2:
            os.remove(check_file_path_2)

        df1 = pd.DataFrame()
        
        #sysinfo_cmd = 'show sysinfo'
        radius_sum_cmd = 'show radius summary'
        wlan_sum_cmd = 'show wlan summary'
        wlan_id = 'show wlan '
        # Create a new temp text files
        # Temp_TXT_1=open(tftp_path + str(ip_address) + str(tempoutput1), "w")
        #Temp_TXT_2=open(tftp_path + str(ip_address) + str(tempoutput2), "w")
        Temp_TXT_3=open(tftp_path + str(ip_address) + str(tempoutput3), "w")
        #Temp_TXT_1=open(tftp_path + str(ip_address) + str(tempoutput1), "a")
        Temp_TXT_2=open(tftp_path + str(tempoutput2), "a")
        #Temp_TXT_3=open(tftp_path + str(tempoutput3), "a")
        # Send the command using Netmiko and store in result
        #result1 = net_connect.send_command(sysinfo_cmd)
        result2 = net_connect.send_command(radius_sum_cmd)
        result3 = net_connect.send_command(wlan_sum_cmd)
        # Write the result to the temp text file and close the file
        #Temp_TXT_1.write(result1)
        Temp_TXT_2.write("Radius Summary for WLC IP Address {}\n".format(ip_address) + result2 + "\n")
        Temp_TXT_3.write(result3)
        #Temp_TXT_1.close()
        Temp_TXT_2.close()
        Temp_TXT_3.close()

        # Get the WLAN IDs configured, clean up the WLAN summary file to only include the 
        # WLAN IDs. Call the clean file function
        clean_file.clean_file(ip_address,tftp_path)
        
        # Load in the cleaned text file
        df1 = pd.read_csv(str(tftp_path) + ip_address + '_Wlan_Summary_clean.csv', header=None)
        
        # Drop any rows that have the following 'Number|WLAN|-' in the row
        df1 = df1.loc[~df1[0].str.contains('Number|WLAN|-', flags=re.I, regex=True)]  
        # Reset the indexes
        df1.reset_index(drop=True, inplace=True)

        for index, row in df1.iterrows():
            df2 = pd.DataFrame()  #  Clear the DataFrame 2
            # Get the summary for the WLAN ID
            int_command = wlan_id + df1.iloc[index,0]
            # Create a new temp text file
            #Temp_TXT=open(tftp_path + ip_address + "_WLAN ID_" + str(df1.iloc[index,0]) + str(tempoutput4), "w")
            Temp_TXT=open(tftp_path + ip_address + str(tempoutput4), "w")
            # Send the command using Netmiko and store in result
            result = net_connect.send_command(int_command)
            # Write the result to the temp text file and close the file
            Temp_TXT.write(result)
            Temp_TXT.close()
            #df2 = pd.read_fwf(tftp_path + ip_address + "_WLAN ID_" + str(df1.iloc[index,0]) + str(tempoutput4), header=None)
            df2 = pd.read_fwf(tftp_path + ip_address + str(tempoutput4), header=None)
            df2 = df2.loc[df2[0].str.contains('SSID|Authentication|Accounting|802.1x', flags=re.I, regex=True, na=False)]
            df2 = df2.loc[~df2[0].str.contains('X|Limits|Broadcast|Web|FlexConnect|Open|EAP|.11', regex=True, na=False)]
            Temp_TXT_5=open(tftp_path + ip_address + str(tempoutput5), "a")
            Temp_TXT_5.close 
            df2.to_csv(tftp_path + ip_address + str(tempoutput5), index=False, header=None, mode='a')
            Temp_TXT_5=open(tftp_path + ip_address + str(tempoutput5), "a")
            Temp_TXT_5.write("\n\n")
            Temp_TXT_5.close 
            #print(df2)

        # Remove any temp files
        os.remove(tftp_path + ip_address + '_Wlan_Summary.txt')
        os.remove(tftp_path + ip_address + '_Wlan_Summary_clean.csv')
        os.remove(tftp_path + ip_address + tempoutput4)
        # Set Task/Thread as complete
        q.task_done()

def main():

    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    # Loop through the IP Address CSV and put the IP address into the queue
    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()

    # Check if there has been a successful connection, if there has remove the 
    # temp sheet. If not then remove the new workbook
    #if 'Success' in connection_details:
        # Remove temp worksheet
    #    workbook.remove(workbook['Sheet1'])
        # Save the workbook
    #    workbook.save(tftp_path_final + spreadsheet_name)
    #else:
    #    os.remove(tftp_path_final + spreadsheet_name)

    print('*** WLC Check Completed ***')

if __name__ == "__main__":
    
    
    # Call the main function
    main()