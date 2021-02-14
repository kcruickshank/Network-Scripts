import pandas as pd
from pandas import ExcelWriter
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading
import clean_int_brief as clean_file
from tkinter import filedialog

# These capture errors relating to hitting ctrl+C (I forget the source)
#signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # IOError: Broken pipe
#signal.signal(signal.SIGINT, signal.SIG_DFL)  # KeyboardInterrupt: Ctrl-C

# Set the number of threads, I've found that 5 is the max
num_threads = 1

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()

# Define some variables to be used later in the script and ask user for some input
#username = input('Enter Username: ')
username = 'local_user'
password = 'M1cr0Lab2003'
file_name = filedialog.askopenfilename(initialdir='C:\Test_Files\', title='Select a file')
#password = getpass.getpass('Enter Password: ')
#file_name = filedialog.askopenfilename(initialdir='C:\TFTP-Root\IP_Helper_Files\Final_Output_Files', title='Select a file')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\IP_Helper_Files\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\IP_Helper_Files\\Final_Output_Files\\'
connection_success = False

# Define the current date and error log file info
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y_%H-%M")  # Set timestamp to current system date
timestamp1 = now.strftime("%d/%m/%Y at %H:%M")  # Set timestamp to current system time including hour and minutes
logfile = timestamp + '_error_connection_log.txt'
tempoutput = '_temp_output.txt'

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)

# Load the workbook based on the selection
workbook = load_workbook(file_name)

# Start to iterate through the IP's in the file

def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()
        print("{}: Acquired IP: {}".format(i,ip_address))

        # Define a switch type
        switch = {
            "device_type": "cisco_ios",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
        }

        # Test the ssh connection and handle any errors and output to text file
        try:
            net_connect = ConnectHandler(**switch)
        except (AuthenticationException):
            with print_lock:
                print("\n{}: ERROR: Authentication failed for {}. Stopping thread. \n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: Authentication failed for {} on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            q.task_done()
            continue 
        except (NetMikoTimeoutException):
            with print_lock:
                print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: Connection to {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            q.task_done()
            continue
        except (SSHException):
            with print_lock:
                print("\n{}: SSH might not be enable on: {} timed-out.\n".format(ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: SSH might not be enabled on: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()
            q.task_done()
            continue
        except (EOFError):
            with print_lock:
                print("\n{}: End of line error attempting device: {} timed-out.\n".format(i,ip_address))
            error_log=open(tftp_path + str(logfile), "a")
            error_log.write("ERROR: End of line error attempting device: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
            error_log.close()    
            q.task_done()
            continue
        
        
        df1 = pd.DataFrame()
        net_connect.enable()
        switchname = net_connect.send_command ("sh ver | i uptime")  # Use to get the hostname
        switchname = switchname.split()[0]  # Get the first word which will be the switchname
        switchname = switchname.strip()  # Strip any trailing white space from the variable
        #net_connect.send_config_set("ip tftp source-interface loopback 0")  # Set tftp source interface to loopback 0

        # Assign IOS commands to variables
        command_1 = "show ip int brief | redirect tftp://" + str(ip_tftp_server) + "/IP_Helper_Files/" + switchname + "_ip_int_brief.txt"
        with print_lock:
            print("Collecting data for " + switchname + "...")
        # Get the text files with the show IP int brief command
        net_connect.send_command(command_1)
        # Clean up the files  
        clean_file.clean_file(switchname,tftp_path)
        # Remove the temp text file   
        os.remove(str(tftp_path) + switchname + '_ip_int_brief.txt')
        # Load in the cleaned text file
        df1 = pd.read_csv(str(tftp_path) + switchname + '_int_brief_clean.csv', header=None)
        # Remove the cleaned file
        os.remove(str(tftp_path) + switchname + '_int_brief_clean.csv')
        # Define column for dataframe
        desc_columns = ['VLAN Interface']
        # Apply the column to the dataframe
        df1.columns = desc_columns
        # Drop any rows that don't have the word VLAN in the row
        df1 = df1.loc[df1['VLAN Interface'].str.contains('Vlan', flags=re.I, regex=True)]  
        # Reset the indexes
        df1.reset_index(drop=True, inplace=True)  
        
        
        for index, row in df1.iterrows():
            df2 = pd.DataFrame()
            # Create new command get interface config
            int_command = 'show run int ' + df1.iloc[index,0]
            # Create a new temp text file
            Temp_TXT=open(tftp_path + str(switchname) + str(tempoutput), "w")
            # Send the command using Netmiko and store in result
            result = net_connect.send_command(int_command)
            # Write the result to the temp text file and close the file
            Temp_TXT.write(result)
            Temp_TXT.close()
            # Read the new Text file into a new dataframe
            df2 = pd.read_csv(str(tftp_path) + str(switchname) + tempoutput, header=None)
            # Drop all rows in the dataframe that don't contain text helper
            df2 = df2.loc[df2[0].str.contains('helper', flags=re.I, regex=True)]
            # Remove text ip helper-address, so it leaves the ip address
            df2 = df2.replace({'ip helper-address': ''}, regex=True)

            
            # If there are no helpers found from the interface
            # then do nothing else remove the helpers found
            if df2.empty:
                 continue
            else:
                df2_list = []  # Create List Variable
                df2_list = df2[0].tolist()  # If DF2 not empty then create a list of IPs
                remove_command = []
                interface = 'interface ' + str(df1.iloc[index,0]).strip()
                remove_command.append(interface)
                for val in df2_list:
                    remove_command.append('no ip helper-address ' + str(val).strip())

            with print_lock:
                print("Removing Helpers on switch " + str(switchname))    
                print(remove_command)
            net_connect.send_config_set(remove_command)

        # Set the active worksheet to current switchname
        worksheet=workbook[switchname]

        # Set Min and Max Rows
        min_row = 2
        max_row = worksheet.max_row
        
        rows = worksheet.iter_rows(min_row,max_row)
        for row in rows:
            curr_col = 1
            add_command = []
            for cell in row:
                if curr_col == 1:
                    int_vlan = 'interface ' + str(cell.value).strip()
                    curr_col += 1  
                    add_command.append(int_vlan)
                    continue
                elif curr_col == 2:
                    # Don't check this cell as it the subnet info
                    curr_col += 1
                    continue
                if cell.value == 'No IP Helper Set':
                    add_command = []
                    break
                else:
                    if not cell.value is None:
                        add_command.append('ip helper-address ' + str(cell.value).strip())
            # Check if the command list is empty, only apply config if not empty
            if len(add_command) != 0:
                net_connect.send_config_set(add_command)
                with print_lock:
                    print("Adding Helpers on switch " + str(switchname))
                    print(add_command)

        # Save switch config
        net_connect.save_config()
        # Disconnect from the switch
        net_connect.disconnect

        # Remove any temp files
        os.remove(tftp_path + str(switchname) + tempoutput)

        # Set Task/Thread as complete
        q.task_done()

def main():

    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    # Loop through the IP Address CSV and put the IP address into the queue
    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()

    print('*** IP Helper Updates Completed ***')

if __name__ == "__main__":
    
    # Call the main function
    main()