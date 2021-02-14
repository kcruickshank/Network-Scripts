import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from pandas import DataFrame
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
#import clean_int_brief as clean_brief
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading
import clean_int_brief as clean_file


#export NET_TEXTFSM=C:\Python_Scripts\NET_OPPS\Lib\site-packages\ntc_templates\templates

# These capture errors relating to hitting ctrl+C (I forget the source)
#signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # IOError: Broken pipe
signal.signal(signal.SIGINT, signal.SIG_DFL)  # KeyboardInterrupt: Ctrl-C

# Set the number of threads
num_threads = 5

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()


# Define some variable to be used later in the script and ask user for some input
#username = input('Enter Username: ')
username = 'adm.cruickshank'
password = getpass.getpass('Enter Password: ')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\IP_Helper_Files\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\IP_Helper_Files\\Final_Output_Files\\'

# Define the current date and error log file info
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y_")  # Set timestamp to current system time
logfile = timestamp + '_helper_check_error_log.txt'
tempoutput = '_temp_output.txt'
spreadsheet_name = timestamp + 'IP_Helper_Check.xlsx'

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)
#count_ip = df_read_ip.shape[0]  # Get the number of rows in column 1

# Create Spreadsheet with blank sheet
#writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name, engine='xlsxwriter')
#writer.save()


# Start to iterate through the IP's in the file

def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()
        print("{}: Acquired IP: {}".format(i,ip_address))

        # Define a switch type
        switch = {
            "device_type": "cisco_ios",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
        }

        # Test the ssh connection and handle any errors and output to text file
        try:
            net_connect = ConnectHandler(**switch)
        except (AuthenticationException):
            # Authfailure=open(tftp_path + str(logfile), "a")
            # Authfailure.write('\nAuthentication Failure for IP ' + ip_address)
            # Authfailure.close()
            with print_lock:
                print("\n{}: ERROR: Authenticaftion failed for {}. Stopping thread. \n".format(i,ip_address))
            q.task_done()
            os.kill(os.getpid(), signal.SIGUSR1)
            connect_issue = True
            continue 
        except (NetMikoTimeoutException):
            with print_lock:
                print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue
            # print ('\n' + 'Timeout to device: ' + ip_address)
            # Timeouts==open(tftp_path + str(logfile), "a")
            # Timeouts.write('\nTimeout to device: ' + ip_address)
            # Timeouts.close()
            # connect_issue = True
            # continue
        # except (SSHException):
        #     print ('SSH might not be enabled: ' + ip_address)
        #     SSHException==open(tftp_path + str(logfile), "a")
        #     SSHException.write('\nSSH might not be enabled: ' + ip_address)
        #     SSHException.close()
        #     connect_issue = True
        #     continue 
        # except (EOFError):
        #     print ('\n' + 'End of liner error attempting device: ' + ip_address)
        #     EOFError==open(tftp_path + str(logfile), "a")
        #     EOFError.write('\nEnd of liner error attempting device: ' + ip_address)
        #     EOFError.close()
        #     connect_issue = True
        #     continue

        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        net_connect.enable()
        switchname = net_connect.send_command ("sh ver | i uptime")  # Use to get the hostname
        switchname = switchname.split()[0]  # Get the first word which will be the switchname
        switchname = switchname.strip()  # Strip any trailing white space from the variable
        net_connect.send_config_set("ip tftp source-interface loopback 0")  # Set tftp source interface to loopback 0

        # Assign IOS commands to variables
        command_1 = "show ip int brief | redirect tftp://" + str(ip_tftp_server) + "/IP_Helper_Files/" + switchname + "_ip_int_brief.txt"
        with print_lock:
            print("Collecting data for " + switchname + "...")
        
        # Get the text files with the show IP int brief command
        net_connect.send_command(command_1)
        # Clean up the files   
        clean_file.clean_file(switchname,tftp_path)  
        # Remove the initial text file
        os.remove(str(tftp_path) + switchname + '_ip_int_brief.txt')
        # Load the new cleaned file into the dataframe  
        df1 = pd.read_csv(str(tftp_path) + switchname + '_int_brief_clean.csv', header=None)
        # Remove the cleaned file  
        os.remove(str(tftp_path) + switchname + '_int_brief_clean.csv')
        # Start to clean up the data further in the dataframe
        desc_columns = ['VLAN Interface']
        df1.columns = desc_columns
        # Drop any rows that don't have VLAN 
        df1 = df1.loc[df1['VLAN Interface'].str.contains('Vlan', flags=re.I, regex=True)]  
         # Reset the indexes
        df1.reset_index(drop=True, inplace=True)
        # Write out new file 
        df1.to_csv(str(tftp_path_final) + switchname + '_int_brief_clean.csv', index=False)
        
        # Now loop through the interfaces to send the command
        # sh run int <and the interface from df1>
        for index, row in df1.iterrows():
            int_command = 'show run int ' + df1.iloc[index,0]
            Temp_TXT=open(tftp_path + str(switchname) + '_' + str(df1.iloc[index,0]) + '.txt', "w")
            result = net_connect.send_command(int_command)
            Temp_TXT.write(result)
            Temp_TXT.close()

        # Disconnect from the switch
        net_connect.disconnect
        # Set Task/Thread as complete
        q.task_done()

def main():

    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()


# with print_lock:
#             print('Writing new sheet for ' + str(switchname))
        
#         # Load the workbook
#         workbook = load_workbook(tftp_path_final + spreadsheet_name)
#         # Create new sheet
#         workbook.create_sheet(switchname)
#         # Set the active worksheet to current switchname
#         worksheet=workbook[switchname]
#         # Hide gridlines on sheet 
#         worksheet.sheet_view.showGridLines = False  

#         # Write the headings to the spreadsheet
#         wcell1 = worksheet.cell(1,1)  # Set cell to A1
#         wcell1.value = "Interface"
#         worksheet['A1'].font = Font(bold=True)  # Bold the Text
#         wcell2 = worksheet.cell(1,2)  # Set cell to B1
#         wcell2.value = "IP Helper Info"
#         worksheet['B1'].font = Font(bold=True)  # Bold the Text
        
#         # Set Column Widths
#         worksheet.column_dimensions['A'].width = 18
#         worksheet.column_dimensions['B'].width = 18
#         worksheet.column_dimensions['C'].width = 18
#         worksheet.column_dimensions['D'].width = 18
#         worksheet.column_dimensions['E'].width = 18
#         worksheet.column_dimensions['F'].width = 18
#         worksheet.column_dimensions['G'].width = 18
#         worksheet.column_dimensions['H'].width = 18
#         worksheet.column_dimensions['I'].width = 18
#         worksheet.column_dimensions['J'].width = 18

#         curr_row = 2  # Set the current row to 2 to start writing date from this row
#         curr_col = 1  # Set the starting column at start of loop to 1

#         for index, row in df1.iterrows():
#             df2 = pd.DataFrame()
#             int_command = 'show run int ' + df1.iloc[index,0]
#             Temp_TXT=open(tftp_path + str(switchname) + str(tempoutput), "w")
#             result = net_connect.send_command(int_command)
#             Temp_TXT.write(result)
#             Temp_TXT.close()
#             df2 = pd.read_csv(str(tftp_path) + str(switchname) + tempoutput, header=None)
#             df2 = df2.loc[df2[0].str.contains('helper', flags=re.I, regex=True)]
#             df2 = df2.replace({'ip helper-address': ''}, regex=True)

#             if df2.empty:
#                 new_cell = worksheet.cell(curr_row,curr_col)
#                 new_cell.value = df1.iloc[index,0]
#                 curr_col += 1
#                 new_cell = worksheet.cell(curr_row,curr_col)
#                 new_cell.value = 'No IP Helper Set'
#                 curr_row += 1
#                 curr_col = 1
#             else:
#                 df2_list = []
#                 df2_list = df2[0].tolist()
#                 new_cell = worksheet.cell(curr_row,curr_col)
#                 new_cell.value = df1.iloc[index,0]
#                 curr_col += 1
#                 for val in df2_list:
#                     new_cell = worksheet.cell(curr_row,curr_col)
#                     new_cell.value = val
#                     curr_col += 1
#                 curr_row += 1
#                 curr_col = 1
        
#         # Find the Max Column
#         max_col = worksheet.max_column
#         max_col_letter = get_column_letter(max_col)
#         # Merge the header column from A2 to the Max Col
#         worksheet.merge_cells('B1:' + str(max_col_letter) + '1')
#         worksheet['B1'].alignment = Alignment(horizontal='center')

#         # Save the workbook
#         workbook.save(tftp_path_final + spreadsheet_name)




    # Load the workbook
    #workbook = load_workbook(tftp_path_final + spreadsheet_name)
    # Remove temp worksheet
    #workbook.remove(workbook['Sheet1'])
    # Save the workbook
    #workbook.save(tftp_path_final + spreadsheet_name)
    print('*** IP Helper Program Completed ***')

if __name__ == "__main__":
    
    # Call the main function
    main()