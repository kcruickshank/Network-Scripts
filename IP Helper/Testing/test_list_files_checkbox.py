from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

file_name = filedialog.askopenfilename(initialdir='C:\TFTP-Root\IP_Helper_Files\Final_Output_Files', title='Select a file')

# Load the workbook based on the selection
workbook = load_workbook(file_name)
switchname = 'gb-ams1-001csw001'
# Set the active worksheet to current switchname
worksheet=workbook[switchname]
# Set Min and Max Rows
min_row = 2
max_row = worksheet.max_row

int_vlan = ""
rows = worksheet.iter_rows(min_row,max_row)
for row in rows:
    curr_col = 1
    for cell in row:
        if curr_col == 1:
            int_vlan = 'interface ' + cell.value
            curr_col += 1
            print(int_vlan)
            continue
        if cell.value == 'No IP Helper Set':
            break
        else:
            if not cell.value is None:
                print(str(cell.value).strip())



