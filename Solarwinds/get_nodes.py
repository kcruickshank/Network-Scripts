from orionsdk import SwisClient
import requests
import getpass
import pandas as pd
import edit_node_name
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys
import subprocess
import time
import tkinter as tk
from tkinter import filedialog
import os
from os import path
import re
from re import search

def start_program(program, exit_code=0):
    subprocess.Popen(program)
    sys.exit(exit_code)

def menu():
    print("\n** Get Solarwind Nodes Program **")
    print("\nPress 1 to get nodes and create new spreadsheet")
    print("Press 2 to update Node Name and DNS")
    print("Press 3 to exit the program")
    choice = input("\nEnter Choice: ")
    return choice

def main():
    loop = True

    while loop == True:
        choice = menu()

        if choice == "1":
            authenticated = False
            counter = 0
            npm_server = '10.251.6.63'
            #print('Use your normal Active Directory Account to connect to Solarwinds\n')
            #ip_tftp_server = "10.251.6.35"
            #tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Solarwinds\\'
            # Set the file save location to the user desktop
            #output_file_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            output_file_path = r"C:\Solarwinds"
            file_path = r"C:\Solarwinds\Orion_Output.xlsx"
            verify = False

            if not path.isdir(output_file_path):
                #print(str(output_file_path) + " does not exist")
                os.mkdir(output_file_path)
                #print(str(output_file_path) + " Created!")

            if path.isfile(file_path):
                try:
                    os.rename(file_path,file_path + "_")
                    #print("Access on file \"" + str(file_path) +"\" is available!")
                    os.rename(file_path+"_",file_path)
                except OSError as e:
                    message = "\nCan't create new file as Spreadsheet is Open, please close Orion_Output.xlsx and run again."
                    print(message)
                    input("\nPress Enter key...")
                    os.system('cls')
                    continue
                    #raise SystemExit

            if not verify:
                from requests.packages.urllib3.exceptions import InsecureRequestWarning
                requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

            while not authenticated:
                #username = input('Enter Username: ')
                #password = getpass.getpass('Enter Password: ')
                username = "mapsacc"
                password = "nMUDIXKZ3V"
                swis_npm = SwisClient(npm_server,username,password)
                try:
                    swis_npm.query("SELECT NodeID FROM Orion.Nodes")
                    authenticated = True
                except:
                    if counter == 0:
                        print("Authentication Error, 1st attempt\n")
                        counter += 1
                    elif counter == 1:
                        print("Authentication Error, 2nd attempt\n")
                        counter += 1
                    elif counter == 2:
                        print("Authentication Error, last attempt\n")
                        counter += 1
                    else:
                        print("Authentication Error, please use a correct username and password\n")
                        input("Press any key to exit")
                        raise SystemExit
                    authenticated = False
                
            print("\nGathering information from Solarwinds and Creating New Spreadsheet...\n")
            orion_res = swis_npm.query("SELECT NodeID, Caption, IP_Address, MachineType, IPAddressGUID FROM Orion.Nodes")
            orion_dns = swis_npm.query("SELECT IPAddressN, Name FROM IPAM.DnsRecord")
            #orion_dns_1 = swis_npm.query("SELECT IPAddressN, IPAddress FROM IPAM.IPNode")
            orion_cust_res = swis_npm.query("SELECT NodeID, DeviceType, Site, City, Country FROM Orion.NodesCustomProperties")
            orion_hardware = swis_npm.query("SELECT NodeID, Model, ServiceTag FROM Orion.HardwareHealth.HardwareInfo")
            df1 = pd.DataFrame(orion_res['results'])
            df1.rename(columns= {'Caption':'Existing Name'}, inplace=True)
            df1.rename(columns= {'IP_Address':'IP Address'}, inplace=True)
            df2 = pd.DataFrame(orion_cust_res['results'])
            df3 = pd.DataFrame(orion_hardware['results'])
            df4 = pd.DataFrame(orion_dns['results'])
            df4 = df4.loc[~df4['Name'].str.contains('in-addr.arpa.', flags=re.I, regex=True)]
            df4.rename(columns= {'Name':'DNS'}, inplace=True)
            df1 = pd.merge(df1, df4, how='left', left_on=['IPAddressGUID'], right_on=['IPAddressN'])
            df1 = df1.drop_duplicates(subset=['Existing Name'])  # Remove duplicate columns
            df1 = df1.drop(columns=['IPAddressGUID', 'IPAddressN'], axis=1)  # Drop the unwanted columns
            df1 = df1[['NodeID','Existing Name','IP Address','DNS','MachineType']]  # Re-order the columns
            # Merge the Hardware info and do not drop rows where the NodeID is
            # not found from the hardware results
            df1 = pd.merge(df1, df3, how='left')
            # Merge in the custom node property details
            df1 = pd.merge(df1, df2)
            # Write out the dataframe to a new excel file
            df1.to_excel(str(output_file_path) + r'\Orion_Output.xlsx', index=False)


            workbook = load_workbook(output_file_path + r'\Orion_Output.xlsx')

            # Set the active worksheet to current switchname
            worksheet=workbook['Sheet1']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 38
            worksheet.column_dimensions['E'].width = 31
            worksheet.column_dimensions['F'].width = 22
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 37
            worksheet.column_dimensions['J'].width = 15
            worksheet.column_dimensions['K'].width = 11

            # Add Auto Filter to columns
            worksheet.auto_filter.ref = worksheet.dimensions

            # Freeze the Header Row
            freeze_row = worksheet['A2']
            worksheet.freeze_panes = freeze_row

            # Set Zoom View Level to 85% for worksheet
            worksheet.sheet_view.zoomScale = 85

            # Rename Sheet1
            worksheet.title = 'Orion Nodes'
            #workbook.sheetnames[1] = 'Orion Nodes'


            # Create new sheet as Edit Node Name
            workbook.create_sheet('Edit Node Name')
            # Set the active worksheet to current switchname
            worksheet=workbook['Edit Node Name']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 35

            # Setup some colours for some cells
            heading_green = '88c184'

            # Select Cell A1 and write data and format
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "Existing Name"
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A2 and write data and format
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "IP Address"
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A3 and write data and format
            wcell3 = worksheet.cell(1,3)  # Set cell to B1
            wcell3.value = "Exisiting DNS Name"
            worksheet['C1'].font = Font(bold=True)  # Bold the Text
            worksheet['C1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A4 and write data and format
            wcell4 = worksheet.cell(1,4)  # Set cell to B1
            wcell4.value = "New Node and DNS Name"
            worksheet['D1'].font = Font(bold=True)  # Bold the Text
            worksheet['D1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour

            # Create new sheet Edit Device Type
            workbook.create_sheet('Edit Device Type')
            # Set the active worksheet to current switchname
            worksheet=workbook['Edit Device Type']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35

            # Write the headings to the spreadsheet
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "Node Name"
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "Device Type"
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')

            # for i in range(1,11):
            #     wcell1 = worksheet.cell(i,27)
            #     wcell1.value = 'Test Value ' + str(i)

            # data_val = DataValidation(type="list",formula1='=$AA:$AA')
            # worksheet.add_data_validation(data_val)
            # for i in range(2,100):
            #     data_val.add(worksheet["B" + str(i)])

            # Create new sheet Edit City
            workbook.create_sheet('Edit City')
            # Set the active worksheet to City
            worksheet=workbook['Edit City']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35

            # Write the headings to the spreadsheet
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "Node Name"
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "City"
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')


            # Create new sheet Edit Hostname
            workbook.create_sheet('Edit Hostname')
            # Set the active worksheet to City
            worksheet=workbook['Edit Hostname']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35

            # Write the headings to the spreadsheet
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "IP Address"
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "New Hostname"
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            # Set Heading Colour
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')


            # Save Changes to Workbook
            workbook.save(str(output_file_path) + r'\Orion_Output.xlsx')

            # Print that the program has completed
            print('New Orion_Output.xlsx has been created and saved to ' + str(output_file_path) + '\n')
            input('Press Enter key...')
            os.system('cls')
        elif choice == "2":
            #print("You chose to update the node name and DNS\n")
            #input('Press Enter key...')
            start_program('python', 'edit_node_name.py')
            os.system('cls')
            menu()
        elif choice == "3":
            break
            raise SystemExit
        else:
            print("Invalid answer, accepts only number 1 or 2 or 3\n")
            input('Press Enter key...')
            os.system('cls')

if __name__ == '__main__':

    main()