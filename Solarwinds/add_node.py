import re
import requests
from orionsdk import SwisClient
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog
import warnings
import getpass


def main():
    npm_server = '10.251.6.63'
    username = 'kenneth.cruickshank'
    #username = input('Enter Username: ')
    password = getpass.getpass('Enter Password: ')
    verify = False
    if not verify:
        from requests.packages.urllib3.exceptions import InsecureRequestWarning
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    swis_npm = SwisClient(npm_server,username,password)
    print("Adding SNMP v3C nodes:")
    file = get_file()
    devices = read_file(file)
    custom_update = read_custom(file)
    for device in devices:
        load_npm(swis_npm,device)
    
    update_custom(swis_npm,custom_update)

    print("\nCompleted")

def get_file():
    # Create Dialogue file to select xlsx file
    #user_dir = os.getenv("USERPROFILE")
    user_dir = 'C:\TFTP-Root\Solarwinds'
    root = tk.Tk()  # Creates root window
    root.withdraw()  # Hides the root window so only the file dialog shows
    file_options = {'initialdir':user_dir, 'filetypes':[('Excel Files','.xlsx'),('All FIles','.*')]}
    file_path = filedialog.askopenfilename(**file_options)  
    if not file_path:
        raise SystemExit  # If no file chosen exit the program
    return file_path

def read_file(file):
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    lst_devices = [{
        'Caption':sheet['A'+str(row)].value,
        'IPAddress':sheet['B'+str(row)].value,
        'Community':sheet['C'+str(row)].value,
        'EngineID':1,
        'ObjectSubType':'SNMP',
        'SNMPV3Username':'ROBOOK',
        'SNMPv3AuthKeyIsPwd':True,
        'SNMPv3PrivKey':'bJfJ36Se#GZh1OWZR0QcRIm07',
        'SNMPv3AuthMethod':'SHA1',
        'SNMPv3PrivKeyIsPwd':True,
        'SNMPv3PrivMethod':'AES128',
        'SNMPVersion':3,
        'SNMPV3AuthKey':'sum9decnTjq25y!ZegmAhptI'
        } for row in range(2,sheet.max_row + 1)]

    print('\nFinished getting devices')
    #input('\nPress Enter to continue')
    return lst_devices

def read_custom(file):
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    custom_updates = [{
        'Caption':sheet['A'+str(row)].value,
        'City':sheet['D'+str(row)].value,
        'Country':sheet['E'+str(row)].value,
        'Device_Role':sheet['F'+str(row)].value,
        'Device_Type':sheet['G'+str(row)].value,
        'Site':sheet['H'+str(row)].value
        } for row in range(2,sheet.max_row + 1)]

    print('\nFinished getting devices')
    #input('\nPress Enter to continue')
    return custom_updates

def load_npm(swis_npm,props):
    print('\nAdding Node {} to Solarwinds...'.format(props['IPAddress']),end="")
    results = swis_npm.create('Orion.Nodes',**props)
    
    # extract the nodeID from the result
    nodeid = re.search(r'(\d+)$', results).group(0)

    pollers_enabled = {
        'N.Status.ICMP.Native': True,
        'N.Status.SNMP.Native': False,
        'N.ResponseTime.ICMP.Native': True,
        'N.ResponseTime.SNMP.Native': False,
        'N.Details.SNMP.Generic': True,
        'N.Uptime.SNMP.Generic': True,
        'N.Cpu.SNMP.HrProcessorLoad': True,
        'N.Memory.SNMP.NetSnmpReal': True,
        'N.AssetInventory.Snmp.Generic': True,
        'N.Topology_Layer3.SNMP.ipNetToMedia': False,
        'N.Routing.SNMP.Ipv4CidrRoutingTable': False
    }

    pollers = []
    for k in pollers_enabled:
        pollers.append(
            {
                'PollerType': k,
                'NetObject': 'N:' + nodeid,
                'NetObjectType': 'N',
                'NetObjectID': nodeid,
                'Enabled': pollers_enabled[k]
            }
        )

    for poller in pollers:
        print("\nAdding poller type: {} with status {}... ".format(poller['PollerType'], poller['Enabled']), end="")
        response = swis_npm.create('Orion.Pollers', **poller)
        print("DONE!")
    
def update_custom(swis_npm,props2):
    print("Custom Property Update Test:")
    for item in props2:
        caption = '{}'.format(item['Caption'])
        city = '{}'.format(item['City'])
        country = '{}'.format(item['Country'])
        device_role = '{}'.format(item['Device_Role'])
        device_type = '{}'.format(item['Device_Type'])
        site = '{}'.format(item['Site'])
        results = swis_npm.query(
        "SELECT Uri FROM Orion.Nodes WHERE Caption = @caption",
        caption=caption)  # Get Uri!
        results1 = swis_npm.query(
        "SELECT NodeID FROM Orion.Nodes WHERE Caption = @caption",
        caption=caption)  # Get NodeID!

        uri = results['results'][0]['Uri']
        node_id = results1['results'][0]['NodeID']

        swis_npm.update(uri + '/CustomProperties', City=city, Country=country, DeviceRole=device_role, DeviceType=device_type, Site=site)
        swis_npm.invoke("Cirrus.Nodes", "AddNodeToNCM", node_id)


        obj = swis_npm.read(uri + '/CustomProperties')
        print (obj) 
    
requests.packages.urllib3.disable_warnings()

    
if __name__ == '__main__':

    main()
        