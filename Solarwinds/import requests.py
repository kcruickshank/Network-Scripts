from orionsdk import SwisClient
import requests
import getpass
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys

npm_server = '10.251.6.63'
#username = input('Enter Username: ')
#password = getpass.getpass('Enter Password: ')
username = 'kenneth.cruickshank'
password = '????????'
ip_tftp_server = "10.251.6.35"
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Solarwinds\\'
verify = False
node_id = '2386'
swis_npm =  SwisClient(npm_server, username, password)
requests.packages.urllib3.disable_warnings()

#orion_res = swis_npm.query("SELECT DnsRecordId, Name, Type, Data FROM IPAM.DnsRecord")
orion_res = swis_npm.query("SELECT NodeID, Address, DisplayName FROM IPAM.DnsServer")
orion_res1 = swis_npm.query("SELECT Sysname, IPAddress, NodeID, Alias FROM IPAM.IPNodeReport")

df1 = pd.DataFrame(orion_res['results'])
df2 = pd.DataFrame(orion_res1['results'])
df1.to_excel(str(tftp_path) + 'Orion_Output_Test.xlsx', index=False)
# df2.to_excel(str(tftp_path) + 'Orion_Output_Test2.xlsx', index=False)
print(df2)


#results = swis.query('SELECT SysName FROM IPAM.IPNode WHERE IPAddress=@ip', ip=ip)
#results1 = swis.query(
#"SELECT URI FROM Orion.Nodes WHERE Caption = @caption",
#caption=caption)  # Get NodeID!

#uri = results1['results'][0]['URI']
#print('URI is : ' + str(uri))       

#swis.query
#swis.query("IPAM.SubnetManagement", "GetFirstAvailableIpForGroup", "10.64.254.0", "24", "Hierarchy Group")
#swis.invoke('IPAM.IPAddressManagement', 'AddPtrRecord ', '10.64.254.102', 'cns.muellergroup.com', '10.80.50.45','10.10.in-addr.arpa')
#swis.invoke(uri, 'IPAM.IPAddressManagement', 'AddPtrRecord ', '10.64.254.102', 'cns.muellergroup.com', '10.80.50.45','10.10.in-addr.arpa')
#swis.invoke("IPAM.IPAddressManagement", "RemovePtrRecord", "10.64.254.13", "cns.muellergroup.com", "10.80.50.45","10.10.in-addr.arpa")