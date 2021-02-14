from orionsdk import SwisClient
import requests
import getpass
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys
import subprocess
import time
import tkinter as tk
from tkinter import filedialog
import os
from os import path
import re
from re import search
import warnings
import datetime as datetime

# Define Variables for Solarwinds
#npm_server = '10.251.6.50'
npm_server = 'ar50npm.ads.muellergroup.com'
username = "mapsacc"
password = "nMUDIXKZ3V"
swis_npm = SwisClient(npm_server,username,password)

# Define some Variables for DNS Updates
dns_zone = "cns.muellergroup.com"
dns_server_1 = "10.96.65.100"
dns_server_2 = '10.96.65.101'
dns_server_3 = '10.80.50.45'

# Set the node details variable that will be used to store read file contents
node_details = ''

# Log file and time variables
now = datetime.datetime.now()  # Get current time
timestamp = now.strftime("%d-%m-%Y")  # Set timestamp to current system date
timestamp_1 = now.strftime("- logged at %H:%M")  # Set timestamp to current system date include hours and minutes
dns_updt_logfile = r"C:\Solarwinds\\DNS Update Error Log " + str(timestamp) + ".txt"
dns_del_logfile = r"C:\Solarwinds\\DNS Delete Error Log " + str(timestamp) + ".txt"
custom_updt_logfile = r"C:\Solarwinds\\Node Name Update Error Log " + str(timestamp) + ".txt"
node_details_updt_logfile = r"C:\Solarwinds\\Node Details Update Error Log " + str(timestamp) + ".txt"

# Define the Directory and File Path Variables
dir_path = r"C:\Solarwinds"
file_path = r"C:\Solarwinds\Orion_Output.xlsx"

def menu():
    print("\n** Magical Solarwinds Program **")
    print("\nEnter 1 to get nodes and create new spreadsheet")
    print("Enter 2 to update Node Name and DNS")
    print("Enter 3 to update Node Details")
    print("Enter 0 to exit the program")
    choice = input("\nEnter Choice: ")
    return choice

def main():
    loop = True
    node_details = ''  # Reset the Node Details back to blank
    while loop == True:
        choice = menu()
        #print("\nChoice is : " + str(choice))  #  Used to check Choice output
        if choice == "1":
            authenticated = False
            counter = 0
            verify = False

            # Check if the Directory exists, if not create it.
            if not path.isdir(dir_path):
                #print(str(dir_path) + " does not exist")
                os.mkdir(dir_path)
                

            # Test to check if the file process is open, if it is warn the user to close and run again
            if path.isfile(file_path):
                try:
                    os.rename(file_path,file_path + "_")
                    #print("Access on file \"" + str(file_path) +"\" is available!")
                    os.rename(file_path+"_",file_path)
                except OSError as e:
                    message = "\nCan't create new file as Spreadsheet is Open, please close Orion_Output.xlsx and run again."
                    print(message)
                    input("\nPress Enter key...")
                    os.system('cls')
                    continue
                    #raise SystemExit

            if not verify:
                from requests.packages.urllib3.exceptions import InsecureRequestWarning
                requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

            while not authenticated:
                #username = input('Enter Username: ')
                #password = getpass.getpass('Enter Password: ')
                #username = "mapsacc"
                #password = "nMUDIXKZ3V"
                swis_npm = SwisClient(npm_server,username,password)
                try:
                    swis_npm.query("SELECT NodeID FROM Orion.Nodes")
                    authenticated = True
                except:
                    if counter == 0:
                        print("Authentication Error, 1st attempt\n")
                        counter += 1
                    elif counter == 1:
                        print("Authentication Error, 2nd attempt\n")
                        counter += 1
                    elif counter == 2:
                        print("Authentication Error, last attempt\n")
                        counter += 1
                    else:
                        print("Authentication Error, please use a correct username and password\n")
                        input("Press any key to exit")
                        raise SystemExit
                    authenticated = False
                
            print("\nGathering information from Solarwinds and Creating New Spreadsheet...\n")
            orion_res = swis_npm.query("SELECT NodeID, Caption, IP_Address, MachineType, IPAddressGUID FROM Orion.Nodes")
            orion_dns = swis_npm.query("SELECT IPAddressN, Name FROM IPAM.DnsRecord")
            #orion_dns_1 = swis_npm.query("SELECT IPAddressN, IPAddress FROM IPAM.IPNode")
            orion_cust_res = swis_npm.query("SELECT NodeID, DeviceType, Site, City, Country FROM Orion.NodesCustomProperties")
            orion_hardware = swis_npm.query("SELECT NodeID, Model, ServiceTag FROM Orion.HardwareHealth.HardwareInfo")
            df1 = pd.DataFrame(orion_res['results'])
            df1.rename(columns= {'Caption':'Existing Name'}, inplace=True)
            df1.rename(columns= {'IP_Address':'IP Address'}, inplace=True)
            df2 = pd.DataFrame(orion_cust_res['results'])
            df3 = pd.DataFrame(orion_hardware['results'])
            df4 = pd.DataFrame(orion_dns['results'])
            df4 = df4.loc[~df4['Name'].str.contains('in-addr.arpa.', flags=re.I, regex=True)]
            df4.rename(columns= {'Name':'DNS'}, inplace=True)
            df1 = pd.merge(df1, df4, how='left', left_on=['IPAddressGUID'], right_on=['IPAddressN'])
            df1 = df1.drop_duplicates(subset=['Existing Name'])  # Remove duplicate columns
            df1 = df1.drop(columns=['IPAddressGUID', 'IPAddressN'], axis=1)  # Drop the unwanted columns
            df1 = df1[['NodeID','Existing Name','IP Address','DNS','MachineType']]  # Re-order the columns
            # Merge the Hardware info and do not drop rows where the NodeID is
            # not found from the hardware results
            df1 = pd.merge(df1, df3, how='left')
            # Merge in the custom node property details
            df1 = pd.merge(df1, df2)
            # Write out the dataframe to a new excel file
            df1.to_excel(str(dir_path) + r'\Orion_Output.xlsx', index=False)

            # Load the new workbook to Apply Formatting
            workbook = load_workbook(dir_path + r'\Orion_Output.xlsx')

            # Set the active worksheet to current switchname
            worksheet=workbook['Sheet1']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 38
            worksheet.column_dimensions['E'].width = 31
            worksheet.column_dimensions['F'].width = 22
            worksheet.column_dimensions['G'].width = 15
            worksheet.column_dimensions['H'].width = 15
            worksheet.column_dimensions['I'].width = 37
            worksheet.column_dimensions['J'].width = 15
            worksheet.column_dimensions['K'].width = 11

            # Add Auto Filter to columns
            worksheet.auto_filter.ref = worksheet.dimensions

            # Freeze the Header Row
            freeze_row = worksheet['A2']
            worksheet.freeze_panes = freeze_row

            # Set Zoom View Level to 85% for worksheet
            worksheet.sheet_view.zoomScale = 85

            # Rename Sheet1
            worksheet.title = 'Orion Nodes'

            # Create new sheet as Edit Node Name
            workbook.create_sheet('Edit Node Name')
            # Set the active worksheet to current switchname
            worksheet=workbook['Edit Node Name']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 35

            # Setup some colours for some cells
            heading_green = '88c184'

            # Select Cell A1 and write data and format
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "Existing Name"
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A2 and write data and format
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "IP Address"
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A3 and write data and format
            wcell3 = worksheet.cell(1,3)  # Set cell to B1
            wcell3.value = "Existing DNS Name"
            worksheet['C1'].font = Font(bold=True)  # Bold the Text
            worksheet['C1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell A4 and write data and format
            wcell4 = worksheet.cell(1,4)  # Set cell to B1
            wcell4.value = "New Node and DNS Name"
            worksheet['D1'].font = Font(bold=True)  # Bold the Text
            worksheet['D1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour

            # Create new sheet Edit Device Type
            workbook.create_sheet('Edit Device Details')
            # Set the active worksheet to current switchname
            worksheet=workbook['Edit Device Details']

            # Set Column Widths
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 35
            worksheet.column_dimensions['D'].width = 35

            # Select Cell A1 and write data and format
            wcell1 = worksheet.cell(1,1)  # Set cell to A1
            wcell1.value = "Node Name"  # Cell Value
            worksheet['A1'].font = Font(bold=True)  # Bold the Text
            worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell B1 and write data and format
            wcell2 = worksheet.cell(1,2)  # Set cell to B1
            wcell2.value = "Device Type"  # Cell Value
            worksheet['B1'].font = Font(bold=True)  # Bold the Text
            worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell C1 and write data and format
            wcell3 = worksheet.cell(1,3)  # Set cell to B1
            wcell3.value = "Site"  # Cell Value
            worksheet['C1'].font = Font(bold=True)  # Bold the Text
            worksheet['C1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell D1 and write data and format
            wcell4 = worksheet.cell(1,4)  # Set cell to B1
            wcell4.value = "City"  # Cell Value
            worksheet['D1'].font = Font(bold=True)  # Bold the Text
            worksheet['D1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            # Select Cell E1 and write data and format
            wcell5 = worksheet.cell(1,5)  # Set cell to B1
            wcell5.value = "Country"  # Cell Value
            worksheet['E1'].font = Font(bold=True)  # Bold the Text
            worksheet['E1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')  # Set Heading Colour
            


            # for i in range(1,11):
            #     wcell1 = worksheet.cell(i,27)
            #     wcell1.value = 'Test Value ' + str(i)

            ### Future Code to add dropdown list ###

            # data_val = DataValidation(type="list",formula1='=$AA:$AA')
            # worksheet.add_data_validation(data_val)
            # for i in range(2,100):
            #     data_val.add(worksheet["B" + str(i)])


            # Save Changes to Workbook
            workbook.save(str(dir_path) + r'\Orion_Output.xlsx')

            # Print that the program has completed
            print('New Orion_Output.xlsx has been created and saved to ' + str(dir_path) + '\n')
            input('Press Enter key to return to Menu...')
            os.system('cls')
        elif choice == "2":
            edit_node_name()
            os.system('cls')
        elif choice == "3":
            edit_node_details()
            os.system('cls')
        elif choice == "0":
            raise SystemExit
        else:
            print("Invalid answer, accepts only number 1,2,3 or 4\n")
            input('Press Enter key to return to Menu...')
            os.system('cls')

def edit_node_name():
    authenticated = False
    counter = 0

    # Check if the file exists, if not inform user and get them to run oprion 1
    if not path.isfile(file_path):
        print('\nOrion_Output.xlsx does not exist in ' + str(dir_path))
        print('\nRun option 1 to create New sheet')
        input('\nPress Enter key to return to Menu')
        os.system('cls')
        main()
    
    verify = False
    if not verify:
        from requests.packages.urllib3.exceptions import InsecureRequestWarning
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    
    while not authenticated:
        try:
            swis_npm.query("SELECT NodeID FROM Orion.Nodes")
            authenticated = True
        except:
            if counter == 0:
                print("Authentication Error, 1st attempt\n")
                counter += 1
            elif counter == 1:
                print("Authentication Error, 2nd attempt\n")
                counter += 1
            elif counter == 2:
                print("Authentication Error, last attempt\n")
                counter += 1
            else:
                print("Authentication Error, please use a correct username and password\n")
                input("\nPress Enter key to return to Menu")
            authenticated = False
    
    print("\n*** Node Name and DNS Update Script Started ***")

    file = file_path
    node_details = read_file_name(file)

    if len(node_details) == 0:
        print("\nThere are no details to update")
        input('\nPress Enter key to return to Menu')
        os.system('cls')
        main()
    else:
        custom_code = update_node(swis_npm,node_details)
        dns_delete_code = delete_dns(swis_npm,node_details)
        dns_update_code = update_dns(swis_npm,node_details)

        # Print a final Summary of the program output
        print("\n*** Summary of Node Name and DNS Updates ***")

        # Check the Custom Attributes Update Code
        if custom_code == 1:
            print("\n- There were errors updating node name, please review error log file {}".format(custom_updt_logfile))
        else:
            print("\n- Node name updates completed with no errors.")

        # Check the DNS Delete Code
        if dns_delete_code == 1:
            print("- There were errors deleting some DNS entries, please review error log file {}".format(dns_del_logfile))
        else:
            print("- DNS deletions completed with no errors.")

        # Check the DNS Update Code
        if dns_update_code == 1:
            print("- There were errors updating some DNS entries, please review error log file {}".format(dns_updt_logfile))
        else:
            print("- DNS updates completed with no errors.")

        input('\nPress Enter key to return to Menu')
        os.system('cls')
        main()

def edit_node_details():
    authenticated = False
    counter = 0

    # Check if the file exists, if not inform user and get them to run oprion 1
    if not path.isfile(file_path):
        print('\nOrion_Output.xlsx does not exist in ' + str(dir_path))
        print('\nRun option 1 to create New sheet')
        input('\nPress Enter key to return to Menu')
        os.system('cls')
        main()

    verify = False
    if not verify:
        from requests.packages.urllib3.exceptions import InsecureRequestWarning
        requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    
    while not authenticated:
        swis_npm = SwisClient(npm_server,username,password)
        try:
            swis_npm.query("SELECT NodeID FROM Orion.Nodes")
            authenticated = True
        except:
            if counter == 0:
                print("Authentication Error, 1st attempt\n")
                counter += 1
            elif counter == 1:
                print("Authentication Error, 2nd attempt\n")
                counter += 1
            elif counter == 2:
                print("Authentication Error, last attempt\n")
                counter += 1
            else:
                print("Authentication Error, please use a correct username and password\n")
                input("\nPress Enter key to return to Menu")
                #raise SystemExit
            authenticated = False
    
    print("\n*** Node Details Update Script Started ***")

    file = file_path
    node_details = read_file_details(file)

    if len(node_details) == 0:
        print("\nThere are no details to update")
        input('\nPress Enter key to return to Menu')
        os.system('cls')
        main()
    else:
        node_details_update_code = update_node_details(swis_npm,node_details)

        # Print a final Summary of the program output
        print("\n*** Summary of Node Details Update ***")

        # Check the Custom Attributes Update Code
        if node_details_update_code == 0:
            print("\n- Script completed with no errors.")
        elif node_details_update_code == 1:
            print("\n- There were errors updating node details, please review error log file {}".format(node_details_updt_logfile))


        input('\nPress Enter key to return to Menu')
        os.system('cls')

def read_file_details(file):
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(file)
    sheet = wb['Edit Device Details']
    lst_updates = [{
        'Caption':sheet['A'+str(row)].value,
        'device_type':sheet['B'+str(row)].value,
        'device_site':sheet['C'+str(row)].value,
        'device_city':sheet['D'+str(row)].value,
        'device_country':sheet['E'+str(row)].value
        } for row in range(2,sheet.max_row + 1)]
    return lst_updates

def update_node_details(swis_npm,props):
    node_details_update_code = 0
    print("\nStarting Solarwind Node Details Update...\n")
    for item in props:
        node_name = '{}'.format(item['Caption'])
        new_device_type = '{}'.format(item['device_type'])
        new_device_site = '{}'.format(item['device_site'])
        new_device_city = '{}'.format(item['device_city'])
        new_device_country = '{}'.format(item['device_country'])
        
        # Check to make sure there is an old and a new value to update node name
        if node_name == 'None':
            error_log=open(str(node_details_updt_logfile), "a")
            error_log.write("ERROR: Node name missing, for update details: ({}), ({}), ({}), ({}) ".format/
            (new_device_type,new_device_site,new_device_city,new_device_country) + str(timestamp_1) + "\n")
            error_log.close()
            node_details_update_code = 1
            continue
        else:
            try:
                # Get the URI of the node to be updated from the Node name
                results = swis_npm.query(
                    "SELECT Uri FROM Orion.Nodes WHERE Caption = @caption",
                     caption=node_name)  # Get Uri!
                
                # Set the URI based on Node Name
                uri = results['results'][0]['Uri']
                
                # Now start to update the certain details of the node depending if they are 
                # not empty value
                if new_device_type == 'None':
                    print("- Node {} has no Device Type details to update.".format(node_name))
                    pass
                else:
                    swis_npm.update(uri + '/CustomProperties', DeviceType=new_device_type)
                    print("- Device Type for {} will be updated to {} ".format(node_name,new_device_type))

                if new_device_site == 'None':
                    print("- Node {} has no Site details to update.".format(node_name))
                    pass
                else:
                    swis_npm.update(uri + '/CustomProperties', Site=new_device_site)
                    print("- Site for node {} will be updated to {} ".format(node_name,new_device_site))

                if new_device_city == 'None':
                    print("- Node {} has no City details to update.".format(node_name))
                    pass
                else:
                    swis_npm.update(uri + '/CustomProperties', City=new_device_city)
                    print("- City for node {} will be updated to {} ".format(node_name,new_device_city))

                if new_device_country == 'None':
                    print("- Node {} has no Country details to update.".format(node_name))
                    pass
                else:
                    swis_npm.update(uri + '/CustomProperties', Country=new_device_country)
                    print("- Country for node {} will be updated to {} ".format(node_name,new_device_country))

            except IndexError:
                error_log=open(str(node_details_updt_logfile), "a")
                error_log.write("ERROR: The node {} trying to be updated does not exist in Solarwinds ".format(node_name) + str(timestamp_1) + "\n")
                error_log.close()
                node_details_update_code = 1
                continue   
    return node_details_update_code

def read_file_name(file):
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(file)
    sheet = wb['Edit Node Name']
    lst_updates = [{
        'Caption':sheet['A'+str(row)].value,
        'ip_add':sheet['B'+str(row)].value,
        'dns_name':sheet['C'+str(row)].value,
        'new_name':sheet['D'+str(row)].value
        } for row in range(2,sheet.max_row + 1)]
    return lst_updates

def update_node(swis_npm,props):
    custom_code = 0
    print("\nStarting Solarwind Node Name Updates...")
    for item in props:
        old_name = '{}'.format(item['Caption'])
        new_name = '{}'.format(item['new_name'])
        ip_add = '{}'.format(item['ip_add'])
        
        # Check to make sure there is an old and a new value to update node name
        if old_name == 'None':
            error_log=open(str(custom_updt_logfile), "a")
            error_log.write("ERROR: Existing node name not set in spreadsheet for IP Address {} ".format(ip_add) + str(timestamp_1) + "\n")
            error_log.close()
            custom_code = 1
            continue
        elif new_name == 'None':
            error_log=open(str(custom_updt_logfile), "a")
            error_log.write("ERROR: New node name not set in spreadsheet for IP Address {} ".format(ip_add) + str(timestamp_1) + "\n")
            error_log.close()
            custom_code = 1
            continue
        else:
            try:
                # Start by updating the Node Name
                # Look up the node based on the name in the existing name column
                results = swis_npm.query(
                    "SELECT Uri FROM Orion.Nodes WHERE Caption = @caption",
                     caption=old_name)  # Get Uri!
                
                # Print output of whats happening to the screen
                print('- Updating Node: ' + old_name + ' to new name: ' + new_name)
                uri = results['results'][0]['Uri']
                swis_npm.update(uri, Caption=new_name )         
                
                # Edit the DNS and System Name Fields
                swis_npm.update(uri, DNS=new_name + "." + dns_zone)
                swis_npm.update(uri, SysName=new_name + "." + dns_zone)
            except IndexError:
                error_log=open(str(custom_updt_logfile), "a")
                error_log.write("ERROR: The existing node name {} trying to be updated does not exist in Solarwinds ".format(old_name) + str(timestamp_1) + "\n")
                error_log.close()
                custom_code = 1
                #print('The node {} to be updated does not exist'.format(item['Caption']))
                continue   
    return custom_code

def delete_dns(swis_npm,props):
    dns_delete_code = 0
    print("\nStarting Solarwinds IPAM DNS Deletion...")
    for item in props:
        ip_add = '{}'.format(item['ip_add'])
        old_dns = '{}'.format(item['dns_name'])
        
        # Check to make sure the required values are in the update sheet to update DNS
        if old_dns == 'None':
            error_log=open(str(dns_del_logfile), "a")
            error_log.write("ERROR: DNS Name is blank in Edit Node Name sheet for IP Address {}, cannot delete DNS entry ".format(ip_add) + str(timestamp_1) + "\n")
            error_log.close()
            dns_delete_code = 1
            continue
        elif ip_add == 'None':
            error_log=open(str(dns_del_logfile), "a")
            error_log.write("ERROR: IP Address is blank in Edit Node Name sheet for DNS Name {}, cannot delete DNS entry ".format(old_dns) + str(timestamp_1) + "\n")
            error_log.close()
            dns_delete_code = 1
        else:
            # First try to remove existing DNS Records
            print("- Starting to remove DNS record " + str(old_dns))
            try:
                swis_npm.invoke('IPAM.IPAddressManagement', 'RemoveDnsARecord',old_dns,ip_add,dns_server_1,dns_zone)
                swis_npm.invoke('IPAM.IPAddressManagement', 'RemoveDnsARecord',old_dns,ip_add,dns_server_2,dns_zone)
                swis_npm.invoke('IPAM.IPAddressManagement', 'RemoveDnsARecord',old_dns,ip_add,dns_server_3,dns_zone)
            except requests.exceptions.HTTPError:
                error_log=open(str(dns_del_logfile), "a")
                error_log.write("ERROR: DNS entry {} does not exist for IP Address {} ".format(old_dns,ip_add) + str(timestamp_1) + "\n")
                error_log.close()
                dns_delete_code = 1
                continue

            print("- Finished removing DNS record " + str(old_dns))
    return dns_delete_code

def update_dns(swis_npm,props):
    dns_update_code = 0
    print("\nStarting Solarwinds IPAM DNS Updates...")
    for item in props:
        ip_add = '{}'.format(item['ip_add'])
        new_name = '{}'.format(item['new_name'])
        
        # Check to make sure the required values are in the update sheet to update DNS
        if new_name == 'None':
            error_log=open(str(dns_updt_logfile), "a")
            error_log.write("ERROR: New Name is blank in Edit Node Name sheet for IP Address {}, cannot update DNS entry ".format(ip_add) + str(timestamp_1) + "\n")
            error_log.close()
            dns_update_code = 1
            continue
        elif ip_add == 'None':
            error_log=open(str(dns_updt_logfile), "a")
            error_log.write("ERROR: IP Address is blank in Edit Node Name sheet for New DNS Name {}, cannot update DNS entry ".format(new_name) + str(timestamp_1) + "\n")
            error_log.close()
            dns_update_code = 1
            continue
        else:
            results_dns = swis_npm.query("SELECT Uri, Status FROM IPAM.IPNode WHERE IPAddress = " + "'" + ip_add + "'")
            
            # Check if results are empty and write to error file that IP does not exist
            if len(results_dns['results']) == 0:
                error_log=open(str(dns_updt_logfile), "a")
                error_log.write("ERROR: IP Address {} does not exist in IPAM, cannot update DNS ".format(ip_add) + str(timestamp_1) + "\n")
                error_log.close()
                dns_update_code = 1
                continue
            else:
                uri_dns = results_dns['results'][0]['Uri']
                status = results_dns['results'][0]['Status']

                if status == 2:
                    error_log=open(str(dns_updt_logfile), "a")
                    error_log.write("ERROR: IP Address {} has status available in IPAM, cannot update attributes ".format(ip_add) + str(timestamp_1) + "\n")
                    error_log.close()
                    dns_update_code = 1
                else:  
                    #Add Hostname field in IPAM IP Address
                    swis_npm.update(uri_dns, DnsBackward=new_name + "." + dns_zone)
                    #Try to add new DNS Records
                    print("- Starting to add DNS record " + str(new_name) + "." + str(dns_zone))
                    try:
                        swis_npm.invoke('IPAM.IPAddressManagement', 'AddDnsARecord',new_name,ip_add,dns_server_1,dns_zone)
                        swis_npm.invoke('IPAM.IPAddressManagement', 'AddDnsARecord',new_name,ip_add,dns_server_2,dns_zone)
                        swis_npm.invoke('IPAM.IPAddressManagement', 'AddDnsARecord',new_name,ip_add,dns_server_3,dns_zone)
                    except requests.exceptions.HTTPError:
                        error_log=open(str(dns_updt_logfile), "a")
                        error_log.write("ERROR: Unable to update DNS record for IP Address {} ".format(ip_add) + str(timestamp_1) + "\n")
                        error_log.close()
                        dns_update_code = 1
                        print("Error adding DNS record!")
                        continue
                
                    print("- Finished adding DNS record " + str(new_name) + "." + str(dns_zone))
    
    return dns_update_code

requests.packages.urllib3.disable_warnings()

if __name__ == '__main__':

    main()