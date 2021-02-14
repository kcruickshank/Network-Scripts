import pandas as pd
from pandas import ExcelWriter
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, Rule
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading
import get_interface as clean_file

# These capture errors relating to hitting ctrl+C (I forget the source)
#signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # IOError: Broken pipe
#signal.signal(signal.SIGINT, signal.SIG_DFL)  # KeyboardInterrupt: Ctrl-C

# Set the number of threads, I've found that 5 is the max
num_threads = 30

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()

# Define some variables to be used later in the script and ask user for some input
username = input('Enter Username: ')
#username = 'adm.cruickshank'
password = getpass.getpass('Enter Password: ')
# Get the name of site being audited
site_name = input('Enter the name of the site or location being audited: ')
ip_tftp_server = "10.251.6.35"
tftp_path_files = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Static_Files\\'
tftp_path = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Last_Input_Check\\'
tftp_path_final = "\\" + "\\" + str(ip_tftp_server) + '\TFTP-Root\\Last_Input_Check\\Final_Output_Files\\'

# Define the current date and error log file info
now = datetime.datetime.now()
timestamp = now.strftime("%d-%m-%Y_")  # Set timestamp to current system time
logfile = timestamp + '_helper_check_error_log.txt'
tempoutput = '_temp_output.txt'
spreadsheet_name = timestamp + 'Last_Input_Check_' + str(site_name).strip() + '.xlsx'

# Read the IP addresses from file
df_read_ip = pd.read_csv(str(tftp_path_files) + 'IP_Address_File.csv', header=None)
#count_ip = df_read_ip.shape[0]  # Get the number of rows in column 1

# Create Spreadsheet with blank sheet
writer = pd.ExcelWriter(tftp_path_final + spreadsheet_name, engine='xlsxwriter')
writer.save()
# Load the workbook
workbook = load_workbook(tftp_path_final + spreadsheet_name)

# Start to iterate through the IP's in the file

def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()
        print("{}: Acquired IP: {}".format(i,ip_address))

        # Define a switch type
        switch = {
            "device_type": "cisco_ios",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
        }

        # Test the ssh connection and handle any errors and output to text file
        try:
            net_connect = ConnectHandler(**switch)
        except (AuthenticationException):
            with print_lock:
                print("\n{}: ERROR: Authenticaftion failed for {}. Stopping thread. \n".format(i,ip_address))
            q.task_done()
            continue 
        except (NetMikoTimeoutException):
            with print_lock:
                print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue
        except (SSHException):
            with print_lock:
                print("\n{}: SSH might not be enable on: {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue
        except (EOFError):
            with print_lock:
                print("\n{}: End of liner error attempting device: {} timed-out.\n".format(i,ip_address))
            q.task_done()
            continue

        df1 = pd.DataFrame()
        df2 = pd.DataFrame()
        net_connect.enable()
        switchname = net_connect.send_command ("sh ver | i uptime")  # Use to get the hostname
        switchname = switchname.split()[0]  # Get the first word which will be the switchname
        switchname = switchname.strip()  # Strip any trailing white space from the variable
        net_connect.send_config_set("ip tftp source-interface loopback 0")  # Set tftp source interface to loopback 0

        # Assign IOS commands to variables
        command_1 = "show int status | redirect tftp://" + str(ip_tftp_server) + "/Last_Input_Check/" + switchname + "_int_status.txt"
        with print_lock:
            print("Collecting data for " + switchname + "...")
        # Get the text files with the show IP int brief command
        net_connect.send_command(command_1)
        # Load in the text file
        df1 = pd.read_csv(str(tftp_path) + switchname + '_int_status.txt', header=None)
        # Remove the temp text file   
        os.remove(str(tftp_path) + switchname + '_int_status.txt')
        # Drop any rows that don't have the word VLAN in the row
        df1 = df1.loc[~df1[0].str.contains('connected', flags=re.I, regex=True)]
        #print(df1)
        df1.to_csv(str(tftp_path) + switchname + '_int_status.txt', index=False, header=None)
        df1= pd.DataFrame()
        # Clean up the files  
        clean_file.clean_file(switchname,tftp_path)
        # Remove the temp text file   
        os.remove(str(tftp_path) + switchname + '_int_status.txt')
        # Load in the cleaned text file
        df1 = pd.read_csv(str(tftp_path) + switchname + '_int_status_clean.csv', header=None)
        #print(df1)
        # Remove the cleaned file
        os.remove(str(tftp_path) + switchname + '_int_status_clean.csv')
        # Define column for dataframe
        desc_columns = ['Interface']
        # Apply the column to the dataframe
        df1.columns = desc_columns
        # Drop any rows that don't have the word VLAN in the row
        df1 = df1.loc[~df1['Interface'].str.contains('Po|Lo|Vlan', flags=re.I, regex=True)]  
        # Reset the indexes
        df1.reset_index(drop=True, inplace=True)  
        
        # Print out status update to screen
        with print_lock:
            print('Writing new sheet for ' + str(switchname))
        
        # Create new sheet as name of the switch
        workbook.create_sheet(switchname)
        # Set the active worksheet to current switchname
        worksheet=workbook[switchname]
        # Hide gridlines on sheet 
        worksheet.sheet_view.showGridLines = False  

        # Setup some colours for some cells
        #light_yellow = 'ebeca4'
        #dark_blue = '0c4672'
        heading_green = '88c184'
        #subheading_orange = 'f4c4a4'
        #white_font = 'ffffff'

        # Write the headings to the spreadsheet
        wcell1 = worksheet.cell(1,1)  # Set cell to A1
        wcell1.value = "Interface"
        worksheet['A1'].font = Font(bold=True)  # Bold the Text
        # Set Heading Colour
        worksheet['A1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
        wcell2 = worksheet.cell(1,2)  # Set cell to B1
        wcell2.value = "Last Input Info"
        worksheet['B1'].font = Font(bold=True)  # Bold the Text
        # Set Heading Colour
        worksheet['B1'].fill = PatternFill(fgColor=heading_green, fill_type='solid')
        
        # Set Column Widths
        worksheet.column_dimensions['A'].width = 18
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 25
        

        curr_row = 2  # Set the current row to 2 to start writing data from this row
        curr_col = 1  # Set the starting column to 1 for the start of loop

        for index, row in df1.iterrows():
            df2 = pd.DataFrame()
            # Create new command get interface config
            int_command = 'show int ' + df1.iloc[index,0] + ' | i Last'
            # Create a new temp text file
            Temp_TXT=open(tftp_path + str(switchname) + str(tempoutput), "w")
            # Send the command using Netmiko and store in result
            result = net_connect.send_command(int_command)
            # Write the result to the temp text file and close the file
            Temp_TXT.write(result)
            Temp_TXT.close()
            # Read the new Text file into a new dataframe
            df2 = pd.read_csv(str(tftp_path) + str(switchname) + tempoutput, header=None)
            # Drop all rows in the dataframe that don't contain text 
            df2 = df2.loc[df2[0].str.contains('Last', flags=re.I, regex=True)]
            # Drop rows in dataframe that have counters in the text
            df2 = df2.loc[~df2[0].str.contains('counters', flags=re.I, regex=True)]

            #print(df2)
            # If the Dataframe is empty, write interface and then No IP 
            # Helper set
            if df2.empty:
                # Set the focus of the current cell based on row and column
                new_cell = worksheet.cell(curr_row,curr_col)
                # Get the interface number from df1 and write to cell
                new_cell.value = df1.iloc[index,0]
                # Move to next column
                curr_col += 1
                # Set the focus of the current cell based on row and column
                new_cell = worksheet.cell(curr_row,curr_col)
                # Write text to the cell
                new_cell.value = 'No Information'
                curr_row += 1  # Move to next row
                curr_col = 1  # Set column back to 1
            else:
                df2_0_list = []  # Create List Variable
                df2_1_list = []
                #df2_2_list = []
                df2_list = []
                df2_0_list = df2[0].tolist()  # If DF2 not empty then create a list
                df2_1_list = df2[1].tolist()  # If DF2 not empty then create a list
                #df2_2_list = df2[2].tolist()  # If DF2 not empty then create a list
                df2_list = df2_0_list + df2_1_list
                
                #print(df2_list)
                # Set the focus of the current cell based on row and column
                new_cell = worksheet.cell(curr_row,curr_col)
                new_cell.value = df1.iloc[index,0]
                curr_col += 1
                for val in df2_list:
                    new_cell = worksheet.cell(curr_row,curr_col)
                    new_cell.value = val
                    curr_col += 1
                curr_row += 1
                curr_col = 1
        
        # Set Min and Max Rows
        min_row = 1
        max_row = worksheet.max_row
        
        # Add some borders
        thin_border = Border(
            left=Side(border_style=BORDER_THIN), 
            right=Side(border_style=BORDER_THIN), 
            top=Side(border_style=BORDER_THIN), 
            bottom=Side(border_style=BORDER_THIN)
            )
        
        rows = worksheet.iter_rows(min_row,max_row)
        for row in rows:
            for cell in row:
                cell.border = thin_border

        # Find the Max Column
        max_col = worksheet.max_column
        max_col_letter = get_column_letter(max_col)
        # Merge the header column from A2 to the Max Col
        worksheet.merge_cells('B1:' + str(max_col_letter) + '1')
        worksheet['B1'].alignment = Alignment(horizontal='center')

        # Save the workbook
        with print_lock:
            print('Saving sheet for ' + str(switchname))
        workbook.save(tftp_path_final + spreadsheet_name)

        # Disconnect from the switch
        net_connect.disconnect

        # Remove any temp files
        os.remove(tftp_path + str(switchname) + tempoutput)

        # Set Task/Thread as complete
        q.task_done()

def main():

    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    # Loop through the IP Address CSV and put the IP address into the queue
    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()

    # Remove temp worksheet
    workbook.remove(workbook['Sheet1'])
    # Save the workbook
    workbook.save(tftp_path_final + spreadsheet_name)
    print('*** Last Input/Output Completed ***')

if __name__ == "__main__":
    # Call the main function
    main()