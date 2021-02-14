import pandas as pd
from pandas import ExcelWriter
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import re
from re import search
import os, signal, json
import shutil
import sys
import netmiko
from netmiko.ssh_exception import AuthenticationException, SSHException, NetMikoTimeoutException
from netmiko import ConnectHandler
import getpass
import datetime as datetime
from queue import Queue
from pprint import pprint
import threading
import os
import tkinter as tk
from tkinter import filedialog
import warnings

# Set the number of threads, I've found that 5 is the max
num_threads = 5

# Define the queue
enclosure_queue = Queue()

# Setup a print lock so only one thread prints at the one time
print_lock = threading.Lock()

# Read the IP addresses from file
df_read_ip = pd.read_csv(r'C:\Test_Files\IP_Address_File.csv', header=None)

# Define some variables
username = 'local_user'
password = '**********'

def get_file():
    # Create Dialogue file to select xlsx file
    start_dir = os.getenv("USERPROFILE")
    #start_dir = r'C:\Test_Files'
    root = tk.Tk()  # Creates root window
    root.withdraw()  # Hides the root window so only the file dialog shows
    file_options = {'initialdir':start_dir, 'filetypes':[('Excel Files','.xlsx'),('All FIles','.*')]}
    file_path = filedialog.askopenfilename(**file_options)  
    if not file_path:
        raise SystemExit  # If no file chosen exit the program
    return file_path

def read_file(file):
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet1']
    lst_updates = [{
        'ip_address':sheet['A'+str(row)].value,
        'switch_name':sheet['B'+str(row)].value
        } for row in range(2,sheet.max_row + 1)]
    return lst_updates

file = get_file()
file_contents = read_file(file)


def deviceconnector(i,q):

    # Loop through the IP's
    while True:
        #print("{}: Waiting for IP address...".format(i))
        ip_address = q.get()          
        print("{}: Acquired IP: {}".format(i,ip_address))
        
        # Loop through the update details sheet and
        # add the IP Address to a new list variable
        new_list = []
        for item in file_contents:
            ip_add = '{}'.format(item['ip_address'])
            new_list.append(ip_add)
        
        # Start by checking that the switch IP Address you are 
        # connecting to actually exists in the update details 
        # list. If it does then perform command else quit the thread
        if ip_address in new_list:
            print(str(ip_address) + " is in the list" )

            # Start new for loop to match the         
            for item in file_contents:
                ip_address_check = '{}'.format(item['ip_address'])
                switch_name = '{}'.format(item['switch_name'])

                if ip_address_check == ip_address:
                    with print_lock:
                        print("IP Address in lookup file is " +str(ip_address_check))
                        print("The switch with IP Address " + str(ip_address) + " will be updated with Hostname " + str(switch_name))

                    # Define a switch type
                    switch = {
                        "device_type": "cisco_ios",
                        "ip": ip_address,
                        "username": username,
                        "password": password,
                    }

                    # Test the ssh connection and handle any errors and output to text file
                    try:
                        net_connect = ConnectHandler(**switch)
                    except (AuthenticationException):
                        with print_lock:
                            print("\n{}: ERROR: Authentication failed for {}. Stopping thread. \n".format(i,ip_address))
                        #error_log=open(tftp_path + str(logfile), "a")
                        #error_log.write("ERROR: Authentication failed for {} on the ".format(ip_address) + timestamp1 + "\n")
                        #error_log.close()
                        q.task_done()
                        continue 
                    except (NetMikoTimeoutException):
                        with print_lock:
                            print("\n{}: ERROR: Connection to {} timed-out.\n".format(i,ip_address))
                        #error_log=open(tftp_path + str(logfile), "a")
                        #error_log.write("ERROR: Connection to {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
                        #error_log.close()
                        q.task_done()
                        continue
                    except (SSHException):
                        with print_lock:
                            print("\n{}: SSH might not be enabled on: {} timed-out.\n".format(ip_address))
                        #error_log=open(tftp_path + str(logfile), "a")
                        #error_log.write("ERROR: SSH might not be enabled on: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
                        #error_log.close()
                        q.task_done()
                        continue
                    except (EOFError):
                        with print_lock:
                            print("\n{}: End of line error attempting device: {} timed-out.\n".format(i,ip_address))
                        #error_log=open(tftp_path + str(logfile), "a")
                        #error_log.write("ERROR: End of line error attempting device: {} timed-out on the ".format(ip_address) + timestamp1 + "\n")
                        #error_log.close()    
                        q.task_done()
                        continue

                    net_connect.enable()
                    command_1 = "hostname " + str(switch_name)
                    net_connect.send_config_set(command_1)

                    
                    # Save switch config
                    net_connect.save_config()
                    # Disconnect from the switch
                    net_connect.disconnect

                    # Set Task/Thread as complete
                    q.task_done()

                    # Exit the for loop
                    exit
                else:
                    continue
        else:
            with print_lock:
                        print("Switch IP Address " + str(ip_address) + " not found in update file")
            q.task_done()    

def main():
    # Setup the threads based on the number given above in the variables
    for i in range(num_threads):
        # Create the thread using the device connector as the function, pass in the thread number
        # and the queue object as the parameters
        thread = threading.Thread(target=deviceconnector, args=(i, enclosure_queue,))
        # Set thread up as a background job
        thread.setDaemon(True)
        # Start the thread
        thread.start()

    # Loop through the IP Address CSV and put the IP address into the queue
    for index, row in df_read_ip.iterrows():
        enclosure_queue.put(df_read_ip.iloc[index, 0])

    # Wait for all threads to be completed
    enclosure_queue.join()

    print('*** Completed ***')

if __name__ == "__main__":
    
    # Call the main function
    main()